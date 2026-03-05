from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, CreateView, UpdateView, DetailView, TemplateView
from django.urls import reverse_lazy
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect
from django.db.models import Q
from django.http import JsonResponse
from django.utils import timezone
from apps.organizations.models import *
from .models import Hazard, HazardPhoto, HazardActionItem
from django.utils.safestring import mark_safe  # ADD THIS IMPORT

from django.contrib.auth import get_user_model
import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .utils import generate_hazard_pdf
from django.views import View
from apps.common.image_utils import compress_image

import json
from django.db.models import Count
from django.db.models.functions import TruncMonth
from .forms import HazardForm
from apps.notifications.services import NotificationService

# Make sure all models are imported
from apps.organizations.models import Plant, Zone, Location, SubLocation



User = get_user_model()


class HazardDashboardView(LoginRequiredMixin, TemplateView):
    """Hazard Management Dashboard"""
    template_name = 'hazards/dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get hazards based on user role (This part is already correct)
        if self.request.user.is_superuser or self.request.user.role.name == 'ADMIN': 
            hazards = Hazard.objects.all()
        elif self.request.user.plant: 
            hazards = Hazard.objects.filter(plant=self.request.user.plant)
        else:
            hazards = Hazard.objects.filter(reported_by=self.request.user)
        
        # Statistics (This part is already correct)
        context['total_hazards'] = hazards.count()
        context['open_hazards'] = hazards.exclude(status__in=['RESOLVED', 'CLOSED']).count()
        context['this_month_hazards'] = hazards.filter(
            incident_datetime__month=datetime.date.today().month,
            incident_datetime__year=datetime.date.today().year
        ).count()
        
        # --- THIS IS THE SECTION TO UPDATE ---
        # Match the context variable names to your template (e.g., 'low_risk' instead of 'low_severity')
        context['critical_hazards'] = hazards.filter(severity='critical').count()
        context['low_risk'] = hazards.filter(severity='low').count()          # <-- UPDATED
        context['medium_risk'] = hazards.filter(severity='medium').count()    # <-- UPDATED
        context['high_risk'] = hazards.filter(severity='high').count()        # <-- UPDATED
        

        # Recent hazards (This part is already correct)
        context['recent_hazards'] = hazards.order_by('-incident_datetime')[:10]
        
        return context


class HazardListView(LoginRequiredMixin, ListView):
    """
    List all hazards with filtering.
    This view now includes specific logic to restrict data visibility based on user roles.
    - ADMIN/Superuser can see all hazards.
    - EMPLOYEE can only see hazards they have personally reported.
    - Other roles (like PLANT HEAD) see hazards related to their assigned plant.
    """
    model = Hazard
    template_name = 'hazards/hazard_list.html'
    context_object_name = 'hazards'
    paginate_by = 20

    def get_queryset(self):
        user = self.request.user
        
        queryset = Hazard.objects.select_related('plant', 'location', 'reported_by').order_by('-incident_datetime')

        # Role-based filtering
        if user.is_superuser or (hasattr(user, 'role') and user.role and user.role.name == 'ADMIN'):
            pass
        elif hasattr(user, 'role') and user.role and user.role.name == 'EMPLOYEE':
            queryset = queryset.filter(reported_by=user)
        elif user.plant:
            queryset = queryset.filter(plant=user.plant)
        else:
            queryset = queryset.filter(reported_by=user)

        # Get filter parameters
        search = self.request.GET.get('search', '')
        hazard_type = self.request.GET.get('hazard_type', '')
        risk_level = self.request.GET.get('risk_level', '')
        status = self.request.GET.get('status', '')
        date_from = self.request.GET.get('date_from', '')
        date_to = self.request.GET.get('date_to', '')

        # Apply filters
        if search:
            queryset = queryset.filter(
                Q(report_number__icontains=search) |
                Q(hazard_title__icontains=search)
            )
        if hazard_type:
            queryset = queryset.filter(hazard_type=hazard_type)
        if risk_level:
            queryset = queryset.filter(severity=risk_level)
        if status:
            queryset = queryset.filter(status=status)
        
        if date_from:
            queryset = queryset.filter(incident_datetime__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(incident_datetime__date__lte=date_to)
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add choices for dropdown filters
        context['hazard_types'] = Hazard.HAZARD_TYPE_CHOICES
        context['risk_levels'] = Hazard.SEVERITY_CHOICES
        context['status_choices'] = Hazard.STATUS_CHOICES

        # Retain filter values in the form after submission
        context['search_query'] = self.request.GET.get('search', '')
        context['selected_hazard_type'] = self.request.GET.get('hazard_type', '')
        context['selected_risk_level'] = self.request.GET.get('risk_level', '')
        context['selected_status'] = self.request.GET.get('status', '')

        return context
class HazardCreateView(LoginRequiredMixin, CreateView):
    model = Hazard
    form_class = HazardForm
    template_name = 'hazards/hazard_create.html'
    success_url = reverse_lazy('hazards:hazard_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        context['cancel_url'] = (self.request.GET.get('next') or self.request.META.get('HTTP_REFERER') or '/')
        context['user_assigned_plants'] = user.assigned_plants.filter(is_active=True)
        if context['user_assigned_plants'].count() == 1:
            plant = context['user_assigned_plants'].first()
            context['user_assigned_zones'] = user.assigned_zones.filter(is_active=True, plant=plant)
            if context['user_assigned_zones'].count() == 1:
                zone = context['user_assigned_zones'].first()
                context['user_assigned_locations'] = user.assigned_locations.filter(is_active=True, zone=zone)
                if context['user_assigned_locations'].count() == 1:
                    location = context['user_assigned_locations'].first()
                    context['user_assigned_sublocations'] = user.assigned_sublocations.filter(
                        is_active=True, location=location
                    )

        context['departments'] = Department.objects.filter(is_active=True).order_by('name')
        return context

    def post(self, request, *args, **kwargs):
        """Override post to handle hazard submissions - bypasses form validation"""
        self.object = None
        # Always use handle_multiple_hazards since template uses custom field names
        return self.handle_multiple_hazards(request)
    
    def handle_multiple_hazards(self, request):
        """Handle single or multiple hazard submissions"""
        user = request.user
        hazard_count = int(request.POST.get('hazard_count', 1))
        
        created_hazards = []
        photos_uploaded_total = 0
        
        print(f"\n{'='*80}")
        print(f"🔄 Processing {hazard_count} hazard(s)")
        print(f"{'='*80}\n")
        
        for hazard_index in range(hazard_count):
            print(f"\n--- Processing Hazard #{hazard_index + 1} ---")
            
            # Create new hazard instance
            hazard = Hazard()
            prefix = f'hazard_{hazard_index}_'
            
            # Reporter fields
            hazard.reported_by = user
            hazard.reporter_name = user.get_full_name()
            hazard.reporter_email = user.email
            hazard.reporter_phone = getattr(user, 'phone', '')
            hazard.report_timestamp = timezone.now()
            hazard.report_source = 'web_portal'
            
            # Get hazard-specific fields
            hazard_type = request.POST.get(f'{prefix}hazard_type')
            hazard_category = request.POST.get(f'{prefix}hazard_category')
            severity = request.POST.get(f'{prefix}severity')
            hazard_description = request.POST.get(f'{prefix}hazard_description')
            immediate_action = request.POST.get(f'{prefix}immediate_action', '')
            
            print(f"  Type: {hazard_type}, Category: {hazard_category}, Severity: {severity}")
            
            # Validate required fields
            if not hazard_type or not hazard_category or not severity or not hazard_description:
                messages.error(request, f'Missing required fields for Hazard #{hazard_index + 1}')
                return redirect('hazards:hazard_create')
            
            hazard.hazard_type = hazard_type
            hazard.hazard_category = hazard_category
            hazard.severity = severity
            hazard.hazard_description = hazard_description
            hazard.immediate_action = immediate_action
            
            # Get location fields - try all possible variations
            plant_id = (
                request.POST.get('plant') or 
                request.POST.get('id_plant') or
                request.POST.get(f'{prefix}plant') or
                request.POST.get(f'{prefix}id_plant')
            )
            
            zone_id = (
                request.POST.get('zone') or 
                request.POST.get('id_zone') or
                request.POST.get(f'{prefix}zone') or
                request.POST.get(f'{prefix}id_zone')
            )
            
            location_id = (
                request.POST.get('location') or 
                request.POST.get('id_location') or
                request.POST.get(f'{prefix}location') or
                request.POST.get(f'{prefix}id_location')
            )
            
            sublocation_id = (
                request.POST.get('sublocation') or 
                request.POST.get('id_sublocation') or
                request.POST.get(f'{prefix}sublocation') or
                request.POST.get(f'{prefix}id_sublocation')
            )
            
            # Fallback to user's assigned locations
            if not plant_id and user.plant:
                plant_id = user.plant.id
            if not zone_id and hasattr(user, 'zone') and user.zone:
                zone_id = user.zone.id
            if not location_id and hasattr(user, 'location') and user.location:
                location_id = user.location.id
            
            print(f"  Plant: {plant_id}, Zone: {zone_id}, Location: {location_id}")
            
            # Validate required location fields
            if not plant_id:
                messages.error(request, f'Plant is required for Hazard #{hazard_index + 1}')
                return redirect('hazards:hazard_create')
            
            if not location_id:
                messages.error(request, f'Location is required for Hazard #{hazard_index + 1}')
                return redirect('hazards:hazard_create')
            
            hazard.plant_id = plant_id
            hazard.zone_id = zone_id if zone_id else None
            hazard.location_id = location_id
            hazard.sublocation_id = sublocation_id if sublocation_id else None
            
            # Incident datetime
            incident_datetime_str = request.POST.get('incident_datetime')
            if incident_datetime_str:
                try:
                    hazard.incident_datetime = datetime.datetime.fromisoformat(incident_datetime_str)
                    if timezone.is_naive(hazard.incident_datetime):
                        hazard.incident_datetime = timezone.make_aware(hazard.incident_datetime)
                except:
                    hazard.incident_datetime = timezone.now()
            else:
                hazard.incident_datetime = timezone.now()
            
            # On behalf logic
            behalf_checkbox = request.POST.get(f'{prefix}behalf_checkbox')
            if behalf_checkbox:
                hazard.behalf_person_name = request.POST.get(f'{prefix}behalf_person_name', '')
                behalf_dept_id = request.POST.get(f'{prefix}behalf_person_dept')
                if behalf_dept_id:
                    hazard.behalf_person_dept_id = behalf_dept_id
            
            # Title
            type_display = dict(Hazard.HAZARD_TYPE_CHOICES).get(hazard_type, hazard_type)
            category_display = dict(Hazard.HAZARD_CATEGORIES).get(hazard_category, hazard_category)
            hazard.hazard_title = f"{type_display} - {category_display}"
            
            # Status
            hazard.status = 'REPORTED'
            hazard.approval_status = 'PENDING'
            
            # Deadline
            severity_days = {'low': 30, 'medium': 15, 'high': 7, 'critical': 1}
            base_date = timezone.now().date()
            hazard.action_deadline = base_date + timezone.timedelta(
                days=severity_days.get(severity, 15)
            )
            
            # Save hazard
            try:
                hazard.save()
                print(f"  ✅ Saved with ID: {hazard.id}")
            except Exception as e:
                print(f"  ❌ Save error: {e}")
                import traceback
                traceback.print_exc()
                messages.error(request, f'Error saving Hazard #{hazard_index + 1}: {str(e)}')
                continue
            
            # Generate report number
            today = timezone.now().date()
            plant_code = hazard.plant.code if hazard.plant else 'UNKN'
            count = Hazard.objects.filter(created_at__date=today).count()
            hazard.report_number = f"HAZ-{plant_code}-{today:%Y%m%d}-{count:03d}"
            hazard.save(update_fields=['report_number'])
            print(f"  📋 Report: {hazard.report_number}")
            
            # Handle photos
            photos_uploaded = 0
            photo_count = int(request.POST.get(f'{prefix}photo_count', 1))
            
            for i in range(photo_count + 5):
                photo_key = f'{prefix}photo_{i}'
                if photo_key in request.FILES:
                    try:
                        photo = request.FILES[photo_key]
                        compressed_photo = compress_image(photo)
                        HazardPhoto.objects.create(
                            hazard=hazard,
                            photo=compressed_photo,
                            photo_type='evidence',
                            uploaded_by=user
                        )
                        photos_uploaded += 1
                    except Exception as e:
                        print(f"  Photo error: {e}")
            
            photos_uploaded_total += photos_uploaded
            created_hazards.append(hazard)
            
            # Send notifications
            try:
                from apps.notifications.services import NotificationService
                NotificationService.notify(
                    content_object=hazard,
                    notification_type='HAZARD_REPORTED',
                    module='HAZARD'
                )
            except Exception as e:
                print(f"  Notification error: {e}")
        
        print(f"\n✅ Total created: {len(created_hazards)}")
        
        # Success messages
        if not created_hazards:
            messages.error(request, 'No hazards were created. Please try again.')
            return redirect('hazards:hazard_create')
        
        if len(created_hazards) == 1:
            hazard = created_hazards[0]
            messages.success(
                request,
                mark_safe(
                    f'<strong>✅ Hazard Report Submitted!</strong><br>'
                    f'Report No: {hazard.report_number}<br>'
                    f'Severity: {hazard.get_severity_display()}<br>'
                    f'Photos: {photos_uploaded_total}'
                )
            )
        else:
            report_numbers = ', '.join([h.report_number for h in created_hazards])
            messages.success(
                request,
                mark_safe(
                    f'<strong>✅ {len(created_hazards)} Hazards Submitted!</strong><br>'
                    f'Reports: {report_numbers}<br>'
                    f'Photos: {photos_uploaded_total}'
                )
            )
        
        return redirect(self.success_url)

    
    
class HazardDetailView(LoginRequiredMixin, DetailView):
    """
    Display details of a specific hazard, optimized for performance.
    """
    model = Hazard
    template_name = 'hazards/hazard_detail.html'
    context_object_name = 'hazard'

    def get_queryset(self):
        """
        Optimize the query by pre-fetching related objects to avoid
        multiple database hits in the template.
        """
       
        return Hazard.objects.select_related(
            'plant', 'zone', 'location', 'sublocation',
            'reported_by', 'assigned_to', 'approved_by',
            'behalf_person', 
            'behalf_person_dept'
        ).prefetch_related(
            'photos', 
            'action_items'  # ✅ FIXED: Just prefetch action_items (no responsible_person)
        )

    def get_context_data(self, **kwargs):
        """
        Add the prefetched photos and action items to the context so the
        template can access them directly.
        """
        
        context = super().get_context_data(**kwargs)
        
        hazard = self.get_object()
        context['action_items'] = hazard.action_items.all()
        context['photos'] = hazard.photos.all()
        context['cancel_url'] = (self.request.GET.get('next') or self.request.META.get('HTTP_REFERER') or '/')
        
        return context

class HazardUpdateView(LoginRequiredMixin, UpdateView):
    model = Hazard
    form_class = HazardForm
    template_name = 'hazards/hazard_update.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_success_url(self):
        return reverse_lazy('hazards:hazard_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        context['cancel_url'] = (self.request.GET.get('next') or self.request.META.get('HTTP_REFERER') or '/')
        # Get user's assigned locations for autofill logic
        context['user_assigned_plants'] = user.assigned_plants.filter(is_active=True)
        
        # Get the current hazard
        hazard = self.object
        
        # Get departments for behalf dropdown
        context['departments'] = Department.objects.filter(is_active=True).order_by('name')
        
        # If user has only one assigned plant, show it as readonly
        if context['user_assigned_plants'].count() == 1:
            plant = context['user_assigned_plants'].first()
            context['user_assigned_zones'] = user.assigned_zones.filter(is_active=True, plant=plant)
            if context['user_assigned_zones'].count() == 1:
                zone = context['user_assigned_zones'].first()
                context['user_assigned_locations'] = user.assigned_locations.filter(is_active=True, zone=zone)
                if context['user_assigned_locations'].count() == 1:
                    location = context['user_assigned_locations'].first()
                    context['user_assigned_sublocations'] = user.assigned_sublocations.filter(
                        is_active=True, location=location
                    )
        
        return context

    def form_valid(self, form):
        hazard = form.save(commit=False)
        user = self.request.user

        # Update reporter fields
        hazard.reporter_name = user.get_full_name()
        hazard.reporter_email = user.email
        hazard.reporter_phone = getattr(user, 'phone', '')
        
        # Update title based on type and category
        type_display = dict(Hazard.HAZARD_TYPE_CHOICES).get(hazard.hazard_type, hazard.hazard_type)
        category_display = dict(Hazard.HAZARD_CATEGORIES).get(hazard.hazard_category, hazard.hazard_category)
        hazard.hazard_title = f"{type_display} - {category_display}"
        
        # Update deadline based on severity
        severity_days = {'low': 30, 'medium': 15, 'high': 7, 'critical': 1}
        base_date = hazard.incident_datetime.date() if hazard.incident_datetime else timezone.now().date()
        hazard.action_deadline = base_date + timezone.timedelta(
            days=severity_days.get(hazard.severity, 15)
        )
        
        # Handle behalf logic
        behalf_checkbox = self.request.POST.get('behalf_checkbox')
        if behalf_checkbox:
            hazard.behalf_person_name = self.request.POST.get('behalf_person_name', '')
            behalf_dept_id = self.request.POST.get('behalf_person_dept')
            if behalf_dept_id:
                hazard.behalf_person_dept_id = behalf_dept_id
        else:
            hazard.behalf_person_name = None
            hazard.behalf_person_dept = None
        
        # Save the hazard
        hazard.save()
        print(f"✅ Hazard updated: {hazard.report_number}")
        
        # Handle photo deletion
        for key in self.request.POST:
            if key.startswith('keep_photo_') and self.request.POST[key] == '0':
                photo_id = key.split('_')[-1]
                HazardPhoto.objects.filter(id=photo_id, hazard=hazard).delete()
                print(f"🗑️ Deleted photo: {photo_id}")

        # Handle new photo uploads
        photo_index = 0
        while True:
            photo_key = f'photo_{photo_index}'
            if photo_key in self.request.FILES:
                try:
                    photo = self.request.FILES[photo_key]
                    compressed_photo = compress_image(photo)
                    HazardPhoto.objects.create(
                        hazard=hazard,
                        photo=compressed_photo,
                        photo_type='evidence',
                        uploaded_by=user
                    )
                    print(f"📸 Added new photo: {photo_key}")
                except Exception as e:
                    print(f"❌ Error uploading photo {photo_key}: {e}")
                photo_index += 1
            else:
                break
        
        # Success message
        messages.success(
            self.request,
            mark_safe(
                f'<strong>✅ Hazard Report Updated!</strong><br>'
                f'Report No: {hazard.report_number}<br>'
                f'Severity: {hazard.get_severity_display()}'
            )
        )
        
        print(f"\n✅ Update complete for {hazard.report_number}")
        print("="*80 + "\n")
        
        return redirect(self.get_success_url())

    def form_invalid(self, form):
        """Handle form validation errors"""
        for field, errors in form.errors.items():
            print(f"  {field}: {errors}")
        return super().form_invalid(form)


                 
class HazardActionItemCreateView(LoginRequiredMixin, CreateView):
    """
    Create an action item for a specific hazard.
    Handles form submission for creating new HazardActionItem(s),
    including file attachments.
    """
    model = HazardActionItem
    template_name = 'hazards/action_item_create.html'
    fields = []

    def dispatch(self, request, *args, **kwargs):
        """Ensure the hazard exists before proceeding."""
        self.hazard = get_object_or_404(Hazard, pk=self.kwargs['hazard_pk'])
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['hazard'] = self.hazard

        # Calculate auto target date based on severity
        if hasattr(self.hazard, 'get_severity_deadline_days'):
            severity_days = self.hazard.get_severity_deadline_days()
        else:
            severity_map = {'low': 30, 'medium': 15, 'high': 7, 'critical': 1}
            severity_days = severity_map.get(self.hazard.severity, 15)

        auto_target_date = timezone.now().date() + timezone.timedelta(days=severity_days)
        context['auto_target_date'] = auto_target_date.strftime('%Y-%m-%d')
        context['severity_days'] = severity_days
        context['cancel_url'] = (self.request.GET.get('next') or self.request.META.get('HTTP_REFERER') or '/')

        user = self.request.user

        if self.hazard.plant:
            from django.db.models import Q

            plant_users = User.objects.filter(
                Q(plant=self.hazard.plant) | Q(assigned_plants=self.hazard.plant),
                is_active=True,
                is_active_employee=True
            ).exclude(
                id=user.id
            ).distinct().select_related(
                'department', 'role'
            ).order_by('first_name', 'last_name')

            context['plant_users'] = plant_users
            context['plant_name'] = self.hazard.plant.name

            print(f"\n🔍 Plant Users Found: {plant_users.count()}")
            for u in plant_users:
                print(f"  - {u.get_full_name()} ({u.email}) - Plant: {u.plant}")

        else:
            context['plant_users'] = []
            context['plant_name'] = 'Unknown Plant'

        return context

    def post(self, request, *args, **kwargs):
        """Handle the POST request to create new action item(s)."""

        print("\n" + "=" * 80)
        print("🎯 ACTION ITEM CREATION")
        print("=" * 80)

        assignment_type = request.POST.get('assignment_type')
        print(f"Assignment Type: {assignment_type}")

        try:
            action_description = request.POST.get('action_description', '').strip()
            target_date_str = request.POST.get('target_date')
            attachment = request.FILES.get('attachment')

            if not target_date_str:
                messages.error(request, 'Target date is required.')
                return redirect('hazards:action_item_create', hazard_pk=self.hazard.pk)

            target_date = datetime.datetime.strptime(target_date_str, '%Y-%m-%d').date()

            # ============================================================
            # ✅ SELF ASSIGNMENT (Complete immediately + Close hazard)
            # ============================================================
            if assignment_type == 'self':

                if not attachment:
                    messages.error(request, 'An attachment is required to self-assign and close an action.')
                    return redirect('hazards:action_item_create', hazard_pk=self.hazard.pk)

                action_item = HazardActionItem.objects.create(
                    hazard=self.hazard,
                    action_description=action_description,
                    created_by=request.user,
                    is_self_assigned=True,
                    target_date=target_date,
                    responsible_emails=request.user.email,
                    status='COMPLETED',
                    completion_date=timezone.now().date(),
                    completion_remarks='Self-assigned and completed by reporter.',
                    completed_by=request.user,
                    attachment=attachment
                )

                print(f"✅ Self-assigned to: {request.user.email}")
                print(f"💾 Action item ID: {action_item.id}")

                # Close hazard
                self.hazard.status = 'CLOSED'
                self.hazard.save(update_fields=['status'])
                print("🔒 Hazard status updated to: CLOSED")

                message = mark_safe(
                    f'✅ <strong>Action item created and completed!</strong><br>'
                    f'Assigned to: <strong>You ({request.user.email})</strong><br>'
                    f'Hazard status: <strong>CLOSED</strong>'
                )
                messages.success(request, message)

            # ============================================================
            # ✅ FORWARD ASSIGNMENT (Create one per selected user)
            # ============================================================
            else:
                responsible_emails = request.POST.getlist('responsible_emails')
                print(f"📧 Responsible emails from POST: {responsible_emails}")

                if not responsible_emails:
                    messages.error(request, 'Please select at least one user to assign this action to.')
                    return redirect('hazards:action_item_create', hazard_pk=self.hazard.pk)

                created_items = []

                for email in responsible_emails:
                    item = HazardActionItem.objects.create(
                        hazard=self.hazard,
                        action_description=action_description,
                        created_by=request.user,
                        is_self_assigned=False,
                        target_date=target_date,
                        responsible_emails=email,
                        status='PENDING',
                        attachment=attachment
                    )
                    created_items.append(item)

                print(f"💾 {len(created_items)} action item(s) created.")

                # Update hazard status
                self.hazard.status = 'ACTION_ASSIGNED'
                self.hazard.save(update_fields=['status'])
                print("📋 Hazard status updated to: ACTION_ASSIGNED")

                # Send notifications
                try:
                    responsible_users = User.objects.filter(
                        email__in=responsible_emails,
                        is_active=True
                    )

                    NotificationService.notify(
                        content_object=item,
                        notification_type='HAZARD_ACTION_ASSIGNED',
                        module='HAZARD_ACTION',
                        extra_recipients=list(responsible_users)
                    )

                except Exception as e:
                    print(f"⚠️ Notification error: {e}")

                message = mark_safe(
                    f'✅ <strong>{len(created_items)} Action Item(s) Created!</strong><br>'
                    f'Assigned to <strong>{len(created_items)}</strong> user(s)<br>'
                    f'Status: <strong>PENDING</strong><br>'
                    f'Hazard status: <strong>ACTION_ASSIGNED</strong>'
                )
                messages.success(request, message)

            print("=" * 80 + "\n")
            return redirect('hazards:hazard_detail', pk=self.hazard.pk)

        except Exception as e:
            print(f"❌ Error creating action item: {e}")
            import traceback
            traceback.print_exc()
            messages.error(request, f'Error creating action item: {str(e)}')
            return redirect('hazards:action_item_create', hazard_pk=self.hazard.pk)

    def get_success_url(self):
        """Redirect to hazard detail page on success."""
        return reverse_lazy('hazards:hazard_detail', kwargs={'pk': self.hazard.pk})
    

class HazardActionItemUpdateView(LoginRequiredMixin, UpdateView):
    """
    Update an existing action item.
    Handles form submission for updating an action item,
    including replacing file attachments. The attachment is always required.
    """
    model = HazardActionItem
    template_name = 'hazards/action_item_update.html'
    fields = []  # We handle fields manually in the post method.

    def dispatch(self, request, *args, **kwargs):
            self.object = self.get_object()
            self.hazard = self.object.hazard
            return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self,**kwargs):
        context = super().get_context_data(**kwargs)
        context['hazard'] = self.hazard
        context['cancel_url'] = (self.request.GET.get('next') or self.request.META.get('HTTP_REFERER') or '/')
        if self.hazard.action_deadline:
            context['action_deadline_date'] = self.hazard.action_deadline.strftime('%Y-%m-%d')
        return context
    
    def get_success_url(self):
        """
        Redirect to the hazard detail page after a successful update.
        """
        return reverse_lazy('hazards:hazard_detail', kwargs={'pk': self.object.hazard.pk})

    def post(self, request, *args, **kwargs):
        """
        Handle the POST request to update the action item.
        """
        self.object = self.get_object()
        
        try:
            # *** MODIFIED: Validate that an attachment will exist after the update ***
            # A new file is being uploaded, which is always acceptable.
            if 'attachment' in request.FILES:
                self.object.attachment = request.FILES['attachment']
            # No new file is uploaded, AND the existing attachment is marked for removal.
            # This logic prevents removing the file without replacing it.
            elif not self.object.attachment:
                messages.error(request, 'An attachment is required. Please upload a file.')
                return redirect('hazards:action_item_update', pk=self.object.pk)

            # Update standard fields from POST data
            self.object.action_description = request.POST.get('action_description', '').strip()
            self.object.responsible_emails = request.POST.get('responsible_emails', '').strip()
            
            # Update target date
            target_date_str = request.POST.get('target_date')
            if target_date_str:
                self.object.target_date = datetime.datetime.strptime(target_date_str, '%Y-%m-%d').date()
            
            # Update status
            self.object.status = request.POST.get('status', self.object.status)

            # Handle completion details if status is 'COMPLETED'
            if self.object.status == 'COMPLETED':
                completion_date_str = request.POST.get('completion_date')
                if completion_date_str:
                    self.object.completion_date = datetime.datetime.strptime(completion_date_str, '%Y-%m-%d').date()
                else:
                    self.object.completion_date = datetime.date.today() # Default to today if not provided
                
                self.object.completion_remarks = request.POST.get('completion_remarks', '').strip()
            
            self.object.save()
            # self.object.hazard.update_status_from_action_items()
            
            # Prepare success message
            email_count = self.object.get_emails_count()
            messages.success(
                request, 
                mark_safe(
                    f'✅ <strong>Action item updated successfully!</strong><br>'
                    f'Status: <strong>{self.object.get_status_display()}</strong> | '
                    f'Assigned to: <strong>{email_count}</strong> email(s)'
                )
            )
            return redirect(self.get_success_url())
            
        except Exception as e:
            # Log the error for debugging purposes
            print(f"Error updating action item: {e}")
            messages.error(request, f'Error updating action-item: {str(e)}')
            return redirect('hazards:action_item_update', pk=self.object.pk)



class HazardDashboardViews(LoginRequiredMixin, TemplateView):
    """
    Advanced Hazard Management Dashboard with working filters.
    """
    template_name = 'hazards/hazards_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        today = datetime.date.today()

        # 1. Get all filter parameters from the URL
        selected_plant = self.request.GET.get('plant', '')
        selected_zone = self.request.GET.get('zone', '')
        selected_location = self.request.GET.get('location', '')
        selected_sublocation = self.request.GET.get('sublocation', '')
        selected_month = self.request.GET.get('month', '')
        selected_severity = self.request.GET.get('severity', '')
        selected_status = self.request.GET.get('status', '')
        selected_category = self.request.GET.get('category', '')    # <-- NEW
        selected_department = self.request.GET.get('department', '') # <-- NEW

        # 2. Build the base queryset based on user role
        if user.is_superuser or getattr(user, 'role', None) and user.role.name == 'ADMIN':
            base_hazards = Hazard.objects.all()
            all_plants = Plant.objects.filter(is_active=True).order_by('name')
        elif getattr(user, 'plant', None):
            base_hazards = Hazard.objects.filter(plant=user.plant)
            all_plants = Plant.objects.filter(id=user.plant.id)
        else:
            base_hazards = Hazard.objects.filter(reported_by=user)
            all_plants = Plant.objects.none()

        # 3. Calculate top-level stats BEFORE applying any filters.
        context['total_hazards'] = base_hazards.count()
        context['open_hazards'] = base_hazards.exclude(status__in=['RESOLVED', 'CLOSED']).count()
        context['overdue_hazards_count'] = base_hazards.filter(action_deadline__lt=today).exclude(status__in=['RESOLVED', 'CLOSED']).count()
        this_month_total = base_hazards.filter(incident_datetime__year=today.year, incident_datetime__month=today.month).count()

        # 4. Apply filters to a new queryset for charts and lists.
        filtered_hazards = base_hazards
        if selected_plant:
            filtered_hazards = filtered_hazards.filter(plant_id=selected_plant)
        if selected_zone:
            filtered_hazards = filtered_hazards.filter(zone_id=selected_zone)
        if selected_location:
            filtered_hazards = filtered_hazards.filter(location_id=selected_location)
        if selected_sublocation:
            filtered_hazards = filtered_hazards.filter(sublocation_id=selected_sublocation)
        if selected_severity:
            filtered_hazards = filtered_hazards.filter(severity=selected_severity)
        if selected_category: # <-- NEW
            filtered_hazards = filtered_hazards.filter(hazard_category=selected_category)
        if selected_department: # <-- NEW
            # NOTE: Yeh filter maanta hai ki aapke Hazard model mein 'department' naam ki ek ForeignKey hai.
            # Agar aapka structure alag hai (jaise behalf_person_dept), to aapko is line ko adjust karna hoga.
            # Example: .filter(Q(department_id=selected_department) | Q(behalf_person_dept_id=selected_department))
            filtered_hazards = filtered_hazards.filter(department_id=selected_department)
            
        if selected_status:
            if selected_status == 'open':
                filtered_hazards = filtered_hazards.exclude(status__in=['RESOLVED', 'CLOSED'])
            else:
                filtered_hazards = filtered_hazards.filter(status=selected_status)
        
        if selected_month:
            try:
                year, month = map(int, selected_month.split('-'))
                filtered_hazards = filtered_hazards.filter(incident_datetime__year=year, incident_datetime__month=month)
            except (ValueError, TypeError):
                pass
        
        context['this_month_hazards'] = filtered_hazards.count() if selected_month else this_month_total

        # 5. Prepare filter dropdown options
        context['plants'] = all_plants
        
        zone_qs = Zone.objects.filter(is_active=True)
        if selected_plant:
            zone_qs = zone_qs.filter(plant_id=selected_plant)
        context['zones'] = zone_qs.order_by('name')

        location_qs = Location.objects.filter(is_active=True)
        if selected_zone:
            location_qs = location_qs.filter(zone_id=selected_zone)
        elif selected_plant and not selected_zone:
             location_qs = location_qs.filter(zone__plant_id=selected_plant)
        context['locations'] = location_qs.order_by('name')

        sublocation_qs = SubLocation.objects.filter(is_active=True)
        if selected_location:
            sublocation_qs = sublocation_qs.filter(location_id=selected_location)
        elif selected_zone and not selected_location:
             sublocation_qs = sublocation_qs.filter(location__zone_id=selected_zone)
        context['sublocations'] = sublocation_qs.order_by('name')
        
        context['month_options'] = [{
            'value': (today - datetime.timedelta(days=i*30)).strftime('%Y-%m'),
            'label': (today - datetime.timedelta(days=i*30)).strftime('%B %Y')
        } for i in range(12)]

        context['all_departments'] = Department.objects.filter(is_active=True).order_by('name') # <-- NEW
        context['all_categories'] = Hazard.HAZARD_CATEGORIES # <-- NEW

        # Pass selected filter values and names back to the template
        context.update({
            'selected_plant': selected_plant, 'selected_zone': selected_zone,
            'selected_location': selected_location, 'selected_sublocation': selected_sublocation,
            'selected_month': selected_month, 'selected_severity': selected_severity,
            'selected_status': selected_status,
            'selected_category': selected_category, # <-- NEW
            'selected_department': selected_department, # <-- NEW
        })
        try:
            if selected_plant: context['selected_plant_name'] = Plant.objects.get(id=selected_plant).name
            if selected_zone: context['selected_zone_name'] = Zone.objects.get(id=selected_zone).name
            if selected_location: context['selected_location_name'] = Location.objects.get(id=selected_location).name
            if selected_sublocation: context['selected_sublocation_name'] = SubLocation.objects.get(id=selected_sublocation).name
            if selected_department: context['selected_department_name'] = Department.objects.get(id=selected_department).name # <-- NEW
            if selected_category: context['selected_category_name'] = dict(Hazard.HAZARD_CATEGORIES).get(selected_category) # <-- NEW
            if selected_month:
                year, month = map(int, selected_month.split('-'))
                context['selected_month_label'] = datetime.date(year, month, 1).strftime('%B %Y')
        except:
             pass
        context['has_active_filters'] = any(context.get(key) for key in ['selected_plant', 'selected_zone', 'selected_location', 'selected_sublocation', 'selected_month', 'selected_severity', 'selected_status', 'selected_category', 'selected_department'])

        # 6. Prepare data for lists and charts using the FILTERED queryset
        context['recent_hazards'] = filtered_hazards.select_related('plant', 'location').order_by('-incident_datetime')[:10]

        # --- PIE CHART DATA ---
        top_categories_query = filtered_hazards.values('hazard_category').annotate(count=Count('hazard_category')).order_by('-count')[:3]
        category_display_map = dict(Hazard.HAZARD_CATEGORIES)
        
        # Data for Pie Chart JavaScript
        top_category_labels, top_category_data, top_category_values = [], [], []
        for item in top_categories_query:
            top_category_labels.append(category_display_map.get(item['hazard_category'], 'Unknown'))
            top_category_data.append(item['count'])
            top_category_values.append(item['hazard_category'])
            
        context['top_category_labels'] = json.dumps(top_category_labels)
        context['top_category_data'] = json.dumps(top_category_data)
        context['top_category_values'] = json.dumps(top_category_values)
        
        # ✅ *** THE CRITICAL FIX IS HERE ***
        # We must pass the actual data (the queryset) to the template, NOT a boolean.
        context['top_hazard_categories'] = top_categories_query  
        # context['top_hazard_categories'] = top_categories_query.exists() # Check if data exists for chart

        # ... (Monthly Trend, Severity, and Status charts ka code waise hi rahega) ...
        # Monthly Trend ...
        six_months_ago = today - datetime.timedelta(days=180)
        monthly_hazards = filtered_hazards.filter(incident_datetime__gte=six_months_ago).annotate(month=TruncMonth('incident_datetime')).values('month').annotate(count=Count('id')).order_by('month')
        context['monthly_labels'] = json.dumps([item['month'].strftime('%b %Y') for item in monthly_hazards])
        context['monthly_data'] = json.dumps([item['count'] for item in monthly_hazards])

        # Severity Distribution ...
        severity_distribution = filtered_hazards.values('severity').annotate(count=Count('id'))
        severity_dict = {item['severity']: item['count'] for item in severity_distribution}
        severity_labels = [choice[1] for choice in Hazard.SEVERITY_CHOICES]
        severity_values = [choice[0] for choice in Hazard.SEVERITY_CHOICES]
        context['severity_labels'] = json.dumps(severity_labels)
        context['severity_data'] = json.dumps([severity_dict.get(val, 0) for val in severity_values])

        # Status Distribution ...
        status_distribution = filtered_hazards.values('status').annotate(count=Count('id')).order_by('-count')
        status_labels, status_keys, status_data = [], [], []
        status_choices_dict = dict(Hazard.STATUS_CHOICES)
        for item in status_distribution:
            status_labels.append(status_choices_dict.get(item['status'], item['status']))
            status_keys.append(item['status'])
            status_data.append(item['count'])
        context['status_labels'] = json.dumps(status_labels)
        context['status_keys'] = json.dumps(status_keys)
        context['status_data'] = json.dumps(status_data)
        
        department_distribution = filtered_hazards.filter(
            behalf_person_dept__isnull=False  # Filter out hazards with no department
        ).values(
            'behalf_person_dept__name'        # Group by the name of the related department
        ).annotate(
            count=Count('id')
        ).order_by('-count')

        department_labels = [item['behalf_person_dept__name'] for item in department_distribution]
        department_data = [item['count'] for item in department_distribution]
        
        context['department_labels'] = json.dumps(department_labels)
        context['department_data'] = json.dumps(department_data)
        context['department_chart_data'] = department_distribution.exists()

        return context
    
# ==================================================
# AJAX VIEWS for Cascading Dropdowns
# These views must exist to support the dashboard filters.
# ==================================================

class GetZonesForPlantAjaxView(LoginRequiredMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        plant_id = request.GET.get('plant_id')
        if not plant_id: return JsonResponse([], safe=False)
        zones = Zone.objects.filter(plant_id=plant_id, is_active=True).values('id', 'name')
        return JsonResponse(list(zones), safe=False)

class GetLocationsForZoneAjaxView(LoginRequiredMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        zone_id = request.GET.get('zone_id')
        if not zone_id: return JsonResponse([], safe=False)
        locations = Location.objects.filter(zone_id=zone_id, is_active=True).values('id', 'name')
        return JsonResponse(list(locations), safe=False)

# This view was missing from your urls.py but is needed for the functionality
class GetSubLocationsForLocationAjaxView(LoginRequiredMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        location_id = request.GET.get('location_id')
        if not location_id: return JsonResponse([], safe=False)
        sublocations = SubLocation.objects.filter(location_id=location_id, is_active=True).values('id', 'name')
        return JsonResponse(list(sublocations), safe=False)


class GetSubLocationsForLocationAjaxView(LoginRequiredMixin, TemplateView):
    """
    AJAX view to get sublocations for a selected location.
    This is called by the JavaScript on the dashboard page.
    """
    def get(self, request, *args, **kwargs):
        # Get the 'location_id' from the GET parameters of the request.
        location_id = request.GET.get('location_id')
        
        # If no location_id is provided, return an empty JSON array.
        if not location_id: 
            return JsonResponse([], safe=False)
            
        # Filter SubLocation objects that are active and belong to the selected location.
        # .values('id', 'name') ensures we only fetch the data we need.
        sublocations = SubLocation.objects.filter(location_id=location_id, is_active=True).values('id', 'name')
        
        # Return the queryset as a JSON response.
        return JsonResponse(list(sublocations), safe=False)
class ExportHazardsView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        user = self.request.user

        # 1. First, establish the base queryset based on user permissions.
        # This logic now mirrors the Incident export view.
        if user.is_superuser or (hasattr(user, 'role') and user.role and user.role.name == 'ADMIN'):
            queryset = Hazard.objects.all()
        elif getattr(user, 'plant', None):
            queryset = Hazard.objects.filter(plant=user.plant)
        else:
            queryset = Hazard.objects.filter(reported_by=user)

        # 2. Now, apply filters from the URL.
        selected_plant = request.GET.get('plant')
        selected_zone = request.GET.get('zone')
        selected_location = request.GET.get('location')
        selected_sublocation = request.GET.get('sublocation')
        selected_severity = request.GET.get('severity')
        selected_status = request.GET.get('status')
        selected_month = request.GET.get('month')

        # The plant filter from the URL is ONLY applied if the user is an Admin/Superuser.
        if selected_plant and (user.is_superuser or (hasattr(user, 'role') and user.role.name == 'ADMIN')):
            queryset = queryset.filter(plant_id=selected_plant)
        
        # Apply other filters to the permission-scoped queryset.
        if selected_zone:
            queryset = queryset.filter(zone_id=selected_zone)
        if selected_location:
            queryset = queryset.filter(location_id=selected_location)
        if selected_sublocation:
            queryset = queryset.filter(sublocation_id=selected_sublocation)
        if selected_severity:
            queryset = queryset.filter(severity__iexact=selected_severity)
        if selected_status == 'open':
            queryset = queryset.exclude(status__in=['RESOLVED', 'CLOSED'])
        
        if selected_month:
            try:
                year, month = map(int, selected_month.split('-'))
                queryset = queryset.filter(incident_datetime__year=year, incident_datetime__month=month)
            except (ValueError, TypeError):
                pass
        
        # Optimize database queries.
        queryset = queryset.select_related(
            'plant', 'zone', 'location', 'sublocation', 'reported_by'
        )

        # --- Excel generation code starts here ---
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Hazards Report'

        # --- Define Styles ---
        header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        wrap_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # --- Headers ---
        headers = [
            'Report Number', 'Title', 'Type', 'Category', 'Severity', 'Status',
            'Incident Datetime', 'Reported By', 'Reported Date', 'Plant', 'Zone',
            'Location', 'Sub-Location', 'Description', 'Action Deadline'
        ]
        sheet.append(headers)
        
        # Style header row
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # --- Data Population ---
        for hazard in queryset:
            row_data = [
                hazard.report_number,
                hazard.hazard_title,
                hazard.get_hazard_type_display(),
                hazard.get_hazard_category_display(),
                hazard.get_severity_display(),
                hazard.get_status_display(),
                hazard.incident_datetime.strftime('%Y-%m-%d %H:%M') if hazard.incident_datetime else '',
                hazard.reported_by.get_full_name() if hazard.reported_by else 'N/A',
                hazard.created_at.strftime('%Y-%m-%d') if hazard.created_at else '',
                hazard.plant.name if hazard.plant else 'N/A',
                hazard.zone.name if hazard.zone else 'N/A',
                hazard.location.name if hazard.location else 'N/A',
                hazard.sublocation.name if hazard.sublocation else 'N/A',
                hazard.hazard_description,
                hazard.action_deadline.strftime('%Y-%m-%d') if hazard.action_deadline else ''
            ]
            sheet.append(row_data)

        # --- Auto-adjust Column Widths and Apply Wrapping ---
        desc_col_letter = get_column_letter(headers.index('Description') + 1)
        title_col_letter = get_column_letter(headers.index('Title') + 1)

        for col_idx, column_cells in enumerate(sheet.columns, 1):
            column_letter = get_column_letter(col_idx)
            # Set a fixed width for columns that need text wrapping
            if column_letter in [desc_col_letter, title_col_letter]:
                sheet.column_dimensions[column_letter].width = 50
                # Apply wrap text to all cells in the description/title column
                for cell in column_cells:
                    cell.alignment = wrap_alignment
            else:
                # Auto-size other columns
                max_length = 0
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                sheet.column_dimensions[column_letter].width = max_length + 2

        # --- HTTP Response ---
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f"Hazards_Report_{timezone.now().strftime('%Y-%m-%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        workbook.save(response)
        return response
    
    
    

class HazardPDFView(LoginRequiredMixin, View):
    """
    Handles the generation and download of a Hazard report in PDF format.
    """
    def get(self, request, *args, **kwargs):
        """
        Processes the GET request to download the PDF for a specific hazard.
        """
        # Retrieve the primary key of the hazard from the URL.
        hazard_pk = self.kwargs.get('pk')
        
        # Fetch the Hazard object from the database, or return a 404 error if not found.
        # This pre-fetches related objects to optimize database queries.
        hazard = get_object_or_404(
            Hazard.objects.select_related(
                'plant', 'zone', 'location', 'sublocation', 
                'reported_by', 'behalf_person_dept'
            ), 
            pk=hazard_pk
        )
        
        # Call the PDF generation utility function and return its response.
        return generate_hazard_pdf(hazard)
    
    
class HazardApprovalView(LoginRequiredMixin, DetailView):
    """
    Displays hazard summary for approval and handles the POST requests
    for approving or rejecting the hazard report.
    """
    model = Hazard
    template_name = 'hazards/hazard_approval.html'
    context_object_name = 'hazard'

    def get_context_data(self, **kwargs):
        """Adds related action items to the context."""
        context = super().get_context_data(**kwargs)
        # Fetch all action items related to this hazard to display them.
        context['action_items'] = self.object.action_items.all()
        return context

    def post(self, request, *args, **kwargs):
        """Handles the 'approve' and 'reject' form submissions."""
        hazard = self.get_object()

        # Check which button was clicked based on its 'name' attribute in the form.
        if 'approve_action' in request.POST:
            # Handle the approval logic.
            hazard.approval_status = 'APPROVED'
            hazard.approved_by = request.user
            hazard.approved_date = timezone.now()
            hazard.approved_remarks = ""  # Clear any previous remarks.
            
            # ✅ CRITICAL FIX: Check if action items exist
            action_items_exist = hazard.action_items.exists()
            
            if action_items_exist:
                # If action items exist, set status to ACTION_ASSIGNED
                hazard.status = 'ACTION_ASSIGNED'
            else:
                # If no action items, set to APPROVED
                hazard.status = 'APPROVED'
            
            hazard.save()
            
            # ✅ NOW call update to ensure correct status based on action item progress
            # This will only work because status is no longer PENDING_APPROVAL
            hazard.update_status_from_action_items()

            messages.success(request, f"Hazard {hazard.report_number} has been approved and is now active.")

        elif 'reject_action' in request.POST:
            # Handle the rejection logic.
            rejection_remarks = request.POST.get('rejection_remarks', '').strip()
            if not rejection_remarks:
                messages.error(request, "Rejection remarks are required to reject the report.")
                return redirect('hazards:hazard_approve', pk=hazard.pk)

            hazard.status = 'REJECTED'
            hazard.approval_status = 'REJECTED'
            hazard.approved_remarks = rejection_remarks
            hazard.approved_by = None
            hazard.approved_date = None
            hazard.save()
            messages.warning(request, f"Hazard {hazard.report_number} has been rejected.")

        return redirect('hazards:hazard_detail', pk=hazard.pk)
    

# Add these RIGHT AFTER your imports, BEFORE the class-based views

from django.views.decorators.http import require_GET

@require_GET
def get_zones_by_plant(request, plant_id):
    """
    Function-based view for cascading dropdown
    URL: /hazards/api/get-zones/<plant_id>/
    """
    try:
        zones = Zone.objects.filter(
            plant_id=plant_id, 
            is_active=True
        ).values('id', 'name', 'code').order_by('name')
        return JsonResponse(list(zones), safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_GET
def get_locations_by_zone(request, zone_id):
    """
    Function-based view for cascading dropdown
    URL: /hazards/api/get-locations/<zone_id>/
    """
    try:
        locations = Location.objects.filter(
            zone_id=zone_id, 
            is_active=True
        ).values('id', 'name', 'code').order_by('name')
        return JsonResponse(list(locations), safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_GET
def get_sublocations_by_location(request, location_id):
    """
    Function-based view for cascading dropdown
    URL: /hazards/api/get-sublocations/<location_id>/
    """
    try:
        sublocations = SubLocation.objects.filter(
            location_id=location_id,
            is_active=True
        ).values('id', 'name', 'code').order_by('name')
        return JsonResponse(list(sublocations), safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)    
    

class MyActionItemsView(LoginRequiredMixin, ListView):
    """
    Display action items assigned to the logged-in user
    """
    model = HazardActionItem
    template_name = 'hazards/my_action_items.html'
    context_object_name = 'action_items'
    paginate_by = 20

    def get_queryset(self):
        user = self.request.user
        
        # Get action items where user's email is in responsible_emails
        queryset = HazardActionItem.objects.filter(
            responsible_emails__icontains=user.email
        ).select_related(
            'hazard', 
            'hazard__plant', 
            'hazard__location',
            'created_by'
        ).order_by('-created_at')
        
        # Apply filters
        status_filter = self.request.GET.get('status', '')
        severity_filter = self.request.GET.get('severity', '')
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        if severity_filter:
            queryset = queryset.filter(hazard__severity=severity_filter)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        user = self.request.user
        
        # Statistics
        all_my_items = HazardActionItem.objects.filter(
            responsible_emails__icontains=user.email
        )
        
        context['total_assigned'] = all_my_items.count()
        context['pending_count'] = all_my_items.filter(status='PENDING').count()
        context['in_progress_count'] = all_my_items.filter(status='IN_PROGRESS').count()
        context['completed_count'] = all_my_items.filter(status='COMPLETED').count()
        context['overdue_count'] = all_my_items.filter(
            status__in=['PENDING', 'IN_PROGRESS'],
            target_date__lt=timezone.now().date()
        ).count()
        
        # Filter values
        context['selected_status'] = self.request.GET.get('status', '')
        context['selected_severity'] = self.request.GET.get('severity', '')
        
        # Choices for filters
        context['status_choices'] = HazardActionItem.STATUS_CHOICES
        context['severity_choices'] = Hazard.SEVERITY_CHOICES
        
        return context


class ActionItemCompleteView(LoginRequiredMixin, UpdateView):
    """
    Allow assigned users to mark action item as complete
    """
    model = HazardActionItem
    template_name = 'hazards/action_item_complete.html'
    fields = []
    
    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        
        user_email = request.user.email
        if user_email not in self.object.responsible_emails:
            messages.error(request, 'You are not assigned to this action item.')
            return redirect('hazards:my_action_items')
        
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['action_item'] = self.object
        context['hazard'] = self.object.hazard
        return context
    
    def post(self, request, *args, **kwargs):
        action_item = self.get_object()
        
        try:
            completion_remarks = request.POST.get('completion_remarks', '').strip()
            completion_date_str = request.POST.get('completion_date')
            
            if not completion_remarks:
                messages.error(request, 'Completion remarks are required.')
                return redirect('hazards:action_item_complete', pk=action_item.pk)
            
            if completion_date_str:
                completion_date = datetime.datetime.strptime(completion_date_str, '%Y-%m-%d').date()
            else:
                completion_date = timezone.now().date()
            
            # --- UPDATE ACTION ITEM ---
            action_item.status = 'COMPLETED'
            action_item.completion_date = completion_date
            action_item.completion_remarks = completion_remarks
            action_item.completed_by = request.user  # <<< KEY CHANGE: SET THE COMPLETER

            if 'completion_attachment' in request.FILES:
                action_item.attachment = request.FILES['completion_attachment']
            
            action_item.save()
            
            # This will now check if ALL related action items are complete
            action_item.hazard.update_status_from_action_items()
            
            try:
                from apps.notifications.services import NotificationService
                NotificationService.notify(
                    content_object=action_item,
                    notification_type='HAZARD_ACTION_COMPLETED',
                    module='HAZARD_ACTION'
                )
            except Exception as e:
                print(f"Notification error: {e}")
            
            messages.success(
                request,
                mark_safe(
                    f'✅ <strong>Action item marked as completed!</strong><br>'
                    f'Hazard: {action_item.hazard.report_number}'
                )
            )
            
            return redirect('hazards:my_action_items')
            
        except Exception as e:
            print(f"Error completing action item: {e}")
            messages.error(request, f'Error: {str(e)}')
            return redirect('hazards:action_item_complete', pk=action_item.pk)