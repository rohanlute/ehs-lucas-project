
from django.db.models import Count, Q, Sum
from datetime import datetime, date
import calendar
import json
from apps.ENVdata.constants import MONTHS
from apps.accounts.models import Plant
from .models import *
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse




class EnvironmentalDataFetcher:
    """
    Dynamic data fetcher for auto-calculated environmental questions
    """
    
    @classmethod
    def get_data_for_plant_year(cls, plant, year):
        """
        Returns auto-calculated values for all questions dynamically
        Format: {'Question Text': {'January': 5, 'February': 3, ...}}
        """
        from .models import EnvironmentalQuestion
        
        result = {}
        
        # ✅ UPDATED: Include INSPECTION source type
        auto_questions = EnvironmentalQuestion.objects.filter(
            is_active=True,
            source_type__in=['INCIDENT', 'HAZARD', 'INSPECTION']  # ⬅️ ADDED INSPECTION
        )
        
        MONTHS = [
            'January', 'February', 'March', 'April', 'May', 'June',
            'July', 'August', 'September', 'October', 'November', 'December'
        ]
        
        for question in auto_questions:
            month_data = {}
            
            for month_num, month_name in enumerate(MONTHS, start=1):
                count = cls.calculate_question_value(
                    question, plant, month_num, year
                )
                month_data[month_name] = count
            
            result[question.question_text] = month_data
        
        return result
    
    @classmethod
    def calculate_question_value(cls, question, plant, month, year):
        """Calculate value for a specific question, plant, and month"""
        
        if question.source_type == 'INCIDENT':
            from apps.accidents.models import Incident

            queryset = Incident.objects.filter(
                plant=plant,
                incident_date__month=month,
                incident_date__year=year
            )
            
            # ✅ Apply primary filter - incident_type is ForeignKey (use _id)
            if question.filter_field == 'incident_type' and question.filter_value:
                queryset = queryset.filter(incident_type_id=question.filter_value)
            elif question.filter_field == 'status' and question.filter_value:
                queryset = queryset.filter(status=question.filter_value)
            elif question.filter_field == 'plant' and question.filter_value:
                queryset = queryset.filter(plant_id=question.filter_value)
            
            # ✅ Apply secondary filter
            if question.filter_field_2 and question.filter_value_2:
                if question.filter_field_2 == 'incident_type':
                    queryset = queryset.filter(incident_type_id=question.filter_value_2)
                elif question.filter_field_2 == 'status':
                    queryset = queryset.filter(status=question.filter_value_2)
                elif question.filter_field_2 == 'plant':
                    queryset = queryset.filter(plant_id=question.filter_value_2)
            
            return queryset.count()
        
        elif question.source_type == 'HAZARD':
            try:
                from apps.hazards.models import Hazard
                
                # ✅ Use incident_datetime (DateTimeField)
                queryset = Hazard.objects.filter(
                    plant=plant,
                    incident_datetime__year=year,
                    incident_datetime__month=month
                )
                
                # ✅ Apply primary filter - hazard_type is CharField (NO _id)
                if question.filter_field == 'hazard_type' and question.filter_value:
                    queryset = queryset.filter(hazard_type=question.filter_value)
                elif question.filter_field == 'severity' and question.filter_value:
                    queryset = queryset.filter(severity=question.filter_value)
                elif question.filter_field == 'status' and question.filter_value:
                    queryset = queryset.filter(status=question.filter_value)
                elif question.filter_field == 'plant' and question.filter_value:
                    queryset = queryset.filter(plant_id=question.filter_value)
                
                # ✅ Apply secondary filter
                if question.filter_field_2 and question.filter_value_2:
                    if question.filter_field_2 == 'hazard_type':
                        queryset = queryset.filter(hazard_type=question.filter_value_2)
                    elif question.filter_field_2 == 'severity':
                        queryset = queryset.filter(severity=question.filter_value_2)
                    elif question.filter_field_2 == 'status':
                        queryset = queryset.filter(status=question.filter_value_2)
                    elif question.filter_field_2 == 'plant':
                        queryset = queryset.filter(plant_id=question.filter_value_2)
                
                return queryset.count()
            
            except ImportError:
                return 0
            except Exception as e:
                print(f"Error calculating hazard data: {e}")
                import traceback
                traceback.print_exc()
                return 0
        
        # ========================================
        # ⬇️ NEW SECTION: INSPECTION CALCULATION
        # ========================================
        elif question.source_type == 'INSPECTION':
            try:
                from apps.inspections.models import InspectionSchedule, InspectionTemplate
                
                # Base queryset - filter by plant, month, year
                queryset = InspectionSchedule.objects.filter(
                    plant=plant,
                    scheduled_date__month=month,
                    scheduled_date__year=year
                )
                
                # ✅ Apply primary filter
                if question.filter_field and question.filter_value:
                    if question.filter_field == 'template':
                        # Template is ForeignKey, use _id
                        queryset = queryset.filter(template_id=question.filter_value)
                    
                    elif question.filter_field == 'inspection_type':
                        # Filter by template's inspection_type
                        queryset = queryset.filter(template__inspection_type=question.filter_value)
                    
                    elif question.filter_field == 'status':
                        # Status is CharField on InspectionSchedule
                        queryset = queryset.filter(status=question.filter_value)
                    
                    elif question.filter_field == 'plant':
                        # Plant is ForeignKey, use _id
                        queryset = queryset.filter(plant_id=question.filter_value)
                    
                    elif question.filter_field == 'assigned_to':
                        # Assigned to is ForeignKey (User), use _id
                        queryset = queryset.filter(assigned_to_id=question.filter_value)
                
                # ✅ Apply secondary filter (optional)
                if question.filter_field_2 and question.filter_value_2:
                    if question.filter_field_2 == 'template':
                        queryset = queryset.filter(template_id=question.filter_value_2)
                    
                    elif question.filter_field_2 == 'inspection_type':
                        queryset = queryset.filter(template__inspection_type=question.filter_value_2)
                    
                    elif question.filter_field_2 == 'status':
                        queryset = queryset.filter(status=question.filter_value_2)
                    
                    elif question.filter_field_2 == 'plant':
                        queryset = queryset.filter(plant_id=question.filter_value_2)
                    
                    elif question.filter_field_2 == 'assigned_to':
                        queryset = queryset.filter(assigned_to_id=question.filter_value_2)
                
                return queryset.count()
            
            except ImportError:
                # Inspection module not installed
                print("Inspection module not found")
                return 0
            except Exception as e:
                print(f"Error calculating inspection data: {e}")
                import traceback
                traceback.print_exc()
                return 0
        
        # Default return for unknown source types
        return 0
    
def get_all_plants_environmental_data(plants):
    questions = EnvironmentalQuestion.objects.filter(is_active=True).order_by("order")

    all_data = (
        MonthlyIndicatorData.objects
        .filter(plant__in=plants)
        .select_related("plant", "indicator", "unit")
    )

    current_year = datetime.now().year
    plants_data = []

    for plant in plants:
        questions_data = []

        for q in questions:
            month_data = {}
            total = 0
            has_values = False
            unit_name = q.default_unit.name if q.default_unit else "Count"

            for month_db, month_label in MonthlyIndicatorData.MONTH_CHOICES:
                if q.source_type in ['INCIDENT', 'HAZARD', 'INSPECTION']:
                    month_number = list(calendar.month_name).index(month_label)

                    value = EnvironmentalDataFetcher.calculate_question_value(q, plant, month_number, current_year)
                    month_data[month_label] = value

                    if value:
                        total += value
                        has_values = True
                else:
                    entry = all_data.filter(plant=plant,indicator=q,month=month_db).first()

                    if entry and entry.value is not None:
                        value = entry.value
                        month_data[month_label] = value

                        try:
                            total += float(str(value).replace(",", ""))
                            has_values = True
                        except (ValueError, TypeError):
                            pass
                    else:
                        month_data[month_label] = ""

            questions_data.append({
                "question": q.question_text,
                "unit": unit_name,
                "month_data": month_data,
                "annual": f"{total:,.2f}" if has_values else "",
            })

        plants_data.append({
            "plant": plant,
            "questions_data": questions_data,
        })

    return plants_data


def generate_environmental_excel(plants_data):
    MONTH_LABELS = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Environmental Data"

    headers = ["Plant", "Question", "Unit", "Annual"] + MONTH_LABELS
    ws.append(headers)

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    ws.freeze_panes = "A2"

    for plant_data in plants_data:
        plant_name = plant_data["plant"].name

        for q in plant_data["questions_data"]:
            row = [
                plant_name,
                q["question"],
                q["unit"],
                q["annual"],
            ]

            for month in MONTH_LABELS:
                row.append(q["month_data"].get(month, ""))

            ws.append(row)

    for column_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            cell.border = thin_border

            if cell.row > 1:
                if isinstance(cell.value, (int, float)) or (
                    isinstance(cell.value, str) and cell.value.replace(".", "").replace(",", "").isdigit()
                ):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(vertical="center")

            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 3

    return wb