import json

from django.http import JsonResponse
from django.shortcuts import render
from django.utils import timezone
from apps.organizations.models import Plant
from .models import SafetyIndicatorEntry, WasteReportData, WasteSummary, EnvironmentEntry
from .report_structure import LAGGING_INDICATOR, LEADING_INDICATOR, MANUFACTURING_WASTE_STRUCTURE, MANUFACTURING_ENVIRONMENT_STRUCTURE, NON_MANUFACTURING_ENVIRONMENT_STRUCTURE
from django.db.models import Sum
from django.views.decorators.csrf import csrf_exempt

import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, PatternFill


# ===============================
# WASTE REPORT SR 
# ===============================
def prepare_waste_rows(structure):

    rows = []
    section_no = 0
    current_section = None

    for r in structure:
        row = r.copy()

        if row.get("section"):
            section_no += 1
            current_section = row["section"]
            row["section_no"] = section_no
        else:
            row["section_no"] = ""

        rows.append(row)

    return rows

# ===============================
# WASTE REPORT VIEW 
# ===============================

def manufacturing_waste_report(request):

    user_plants = Plant.objects.all()

    plant_id = request.GET.get("plant_id")
    waste_type = request.GET.get("waste_type", "MANUFACTURING")

    if plant_id:
        plant = Plant.objects.get(id=plant_id)
    else:
        plant = user_plants.first()

    year = timezone.now().year

    # ================= MAIN REPORT DATA =================

    report_data = WasteReportData.objects.filter(
        plant=plant,
        year=year,
        report_type=waste_type
    )

    data_map = {d.row_name: d for d in report_data}

    # ================= SUMMARY DATA =================

    summary_data = WasteSummary.objects.filter(
        plant=plant,
        year=year,
        report_type=waste_type
    )

    summary_map = {s.summary_type: s for s in summary_data}


    # ================= PRODUCTION FROM ENVIRONMENT REPORT =================
    # ================= SELECT ENVIRONMENT REPORT TYPE =================

    if waste_type == "MANUFACTURING":
        env_report_type = "MANUFACTURING_ENV"
    else:
        env_report_type = "NON_MANUFACTURING_ENV"


    # ================= PRODUCTION FROM ENVIRONMENT REPORT =================

    env_entries = EnvironmentEntry.objects.filter(
        plant=plant,
        year=year,
        report_type=env_report_type
    )

    prod = env_entries.aggregate(
        jan=Sum("jan_qty"),
        feb=Sum("feb_qty"),
        mar=Sum("mar_qty"),

        apr=Sum("apr_qty"),
        may=Sum("may_qty"),
        jun=Sum("jun_qty"),

        jul=Sum("jul_qty"),
        aug=Sum("aug_qty"),
        sep=Sum("sep_qty"),

        oct=Sum("oct_qty"),
        nov=Sum("nov_qty"),
        dec=Sum("dec_qty"),
    )

    # replace None with 0
    for k,v in prod.items():
        prod[k] = v or 0

    # quarters
    prod["q1"] = prod["jan"] + prod["feb"] + prod["mar"]
    prod["q2"] = prod["apr"] + prod["may"] + prod["jun"]
    prod["q3"] = prod["jul"] + prod["aug"] + prod["sep"]
    prod["q4"] = prod["oct"] + prod["nov"] + prod["dec"]

    # yearly total
    prod["total"] = sum([
        prod["jan"],prod["feb"],prod["mar"],
        prod["apr"],prod["may"],prod["jun"],
        prod["jul"],prod["aug"],prod["sep"],
        prod["oct"],prod["nov"],prod["dec"]
    ])

    # ================= CONTEXT =================

    context = {
        "report_rows": prepare_waste_rows(MANUFACTURING_WASTE_STRUCTURE),
        "data_map": data_map,
        "summary_map": summary_map,   
        "production_summary": prod,
        "selected_plant": plant,
        "user_plants": user_plants,
        "current_year": year,
        "selected_type": waste_type
    }

    return render(
        request,
        "environmental_mis/manufacturing_waste_report.html",
        context
    )


# ===============================
# WASTE REPORT SAVE 
# ===============================

def save_waste_report(request):

    if request.method != "POST":
        return JsonResponse({"status":"error"},status=400)

    data=json.loads(request.body)

    rows=data.get("rows",[])
    summary=data.get("summary",[])

    plant_id = request.GET.get("plant_id")
    waste_type = request.GET.get("type")

    print("waste_type:", waste_type)
    print("plant_id:", plant_id)

    if waste_type == "MANUFACTURING":
        report_type = "MANUFACTURING"
        print(" If report_type:", report_type)
    else:
        report_type = "NON_MANUFACTURING"
        print("else report_type:", report_type)
    plant=Plant.objects.get(id=plant_id)

    year=timezone.now().year


    # ================= MAIN TABLE =================

    for row in rows:

        row_name=row.pop("row_name")

        WasteReportData.objects.update_or_create(

            plant=plant,
            year=year,
            report_type=report_type,
            row_name=row_name,

            defaults=row

        )


    # ================= SUMMARY TABLE =================

    for item in summary:

        summary_type=item.pop("summary_type")

        WasteSummary.objects.update_or_create(

            plant=plant,
            year=year,
            report_type=report_type,
            summary_type=summary_type,

            defaults=item

        )


    return JsonResponse({"status":"success"})


# ===============================
# ENVIRONMENT REPORT SR 
# ===============================

def prepare_rows(structure):

    rows = []
    current_section = ""
    section_no = 0

    for r in structure:

        row = r.copy()

        if row.get("section"):
            current_section = row["section"]
            section_no += 1
            row["sr_no"] = section_no
        else:
            row["sr_no"] = ""

        row["category"] = current_section
        row["subcategory"] = row.get("sub_section", "")
        row["question"] = row.get("question", "")
        row["unit_category"] = row.get("unit_category", "")
        row["unit"] = row.get("unit", "")
        row["row_key"] = row.get("row_key", "")

        rows.append(row)

    # ================= CATEGORY ROWSPAN =================

    category_counts = {}

    for row in rows:
        cat = row["category"]
        category_counts[cat] = category_counts.get(cat, 0) + 1

    shown_category = set()

    for row in rows:

        cat = row["category"]

        if cat not in shown_category:
            row["category_rowspan"] = category_counts[cat]
            row["sr_rowspan"] = category_counts[cat]   # ADD THIS
            shown_category.add(cat)
        else:
            row["category_rowspan"] = None
            row["sr_rowspan"] = None   # ADD THIS
    # ================= SUBCATEGORY ROWSPAN =================

    subcategory_counts = {}

    for row in rows:

        key = (row["category"], row["subcategory"])

        if row["subcategory"]:
            subcategory_counts[key] = subcategory_counts.get(key, 0) + 1


    shown_subcategory = set()

    for row in rows:

        key = (row["category"], row["subcategory"])

        if not row["subcategory"]:
            row["subcategory_rowspan"] = None
            continue

        if key not in shown_subcategory:
            row["subcategory_rowspan"] = subcategory_counts[key]
            shown_subcategory.add(key)
        else:
            row["subcategory_rowspan"] = None

    return rows

# ===============================
# ENVIRONMENT REPORT VIEW 
# ===============================

def environment_report(request):

    user_plants = Plant.objects.all()

    selected_type = request.GET.get("type", "MANUFACTURING")
    plant_id = request.GET.get("plant_id")

    if plant_id:
        plant = Plant.objects.get(id=plant_id)
    else:
        plant = user_plants.first()

    year = timezone.now().year

    manufacturing_rows = prepare_rows(MANUFACTURING_ENVIRONMENT_STRUCTURE)
    non_manufacturing_rows = prepare_rows(NON_MANUFACTURING_ENVIRONMENT_STRUCTURE)

    months = [
        "jan","feb","mar",
        "apr","may","jun",
        "jul","aug","sep",
        "oct","nov","dec"
    ]

    report_type = (
        "MANUFACTURING_ENV"
        if selected_type == "MANUFACTURING"
        else "NON_MANUFACTURING_ENV"
    )

    # ================= LOAD SAVED DATA =================

    saved_entries = EnvironmentEntry.objects.filter(
        plant=plant,
        year=year,
        report_type=report_type
    )

    data_map = {
        entry.row_name: entry
        for entry in saved_entries
    }

    context = {
        "manufacturing_rows": manufacturing_rows,
        "non_manufacturing_rows": non_manufacturing_rows,
        "months": months,
        "selected_type": selected_type,
        "selected_plant": plant,
        "user_plants": user_plants,
        "current_year": year,
        "data_map": data_map
    }

    return render(
        request,
        "environmental_mis/environment_report.html",
        context
    )


# ===============================
# ENVIRONMENT REPORT SAVE 
# ===============================

def save_environment_report(request):

    if request.method != "POST":
        return JsonResponse({"status": "error"})

    data = json.loads(request.body)
    print("data:", data)

    plant_id = data.get("plant_id")
    print("plant_id:", plant_id)
    report_type = data.get("report_type")
    print("report_type:", report_type)

    year = timezone.now().year

    rows = data.get("rows", [])

    plant = Plant.objects.get(id=plant_id)

    if report_type == "MANUFACTURING":
        report_type = "MANUFACTURING_ENV"
    else:
        report_type = "NON_MANUFACTURING_ENV"

    months = [
        "jan_qty","feb_qty","mar_qty","apr_qty","may_qty","jun_qty",
        "jul_qty","aug_qty","sep_qty","oct_qty","nov_qty","dec_qty"
    ]

    for row in rows:

        row_name = row.get("row_name")

        # Convert empty values to 0
        for m in months:
            val = row.get(m)
            row[m] = float(val) if val not in ["", None] else 0

        # Calculate totals
        row["total_quantity"] = sum(row[m] for m in months)

        row["q1_quantity"] = row["jan_qty"] + row["feb_qty"] + row["mar_qty"]
        row["q2_quantity"] = row["apr_qty"] + row["may_qty"] + row["jun_qty"]
        row["q3_quantity"] = row["jul_qty"] + row["aug_qty"] + row["sep_qty"]
        row["q4_quantity"] = row["oct_qty"] + row["nov_qty"] + row["dec_qty"]

        EnvironmentEntry.objects.update_or_create(
            plant=plant,
            year=year,
            report_type=report_type,
            row_name=row_name,
            defaults=row
        )

    return JsonResponse({"status": "success"})


# ===============================
# SAFETY REPORT VIEW 
# ===============================

def safety_indicator(request):

    user_plants = Plant.objects.all()

    selected_type = request.GET.get("type", "LEADING_INDICATOR")
    plant_id = request.GET.get("plant_id")

    if plant_id:
        plant = Plant.objects.get(id=plant_id)
    else:
        plant = user_plants.first()

    year = timezone.now().year

    leading_indicator_rows = prepare_rows(LEADING_INDICATOR)
    lagging_indicator_rows = prepare_rows(LAGGING_INDICATOR)

    months = [
        "jan","feb","mar",
        "apr","may","jun",
        "jul","aug","sep",
        "oct","nov","dec"
    ]

    report_type = (
        "LEADING_INDICATOR"
        if selected_type == "LEADING_INDICATOR"
        else "LAGGING_INDICATOR"
    )

    # ================= LOAD SAVED DATA =================

    saved_entries = SafetyIndicatorEntry.objects.filter(
        plant=plant,
        year=year,
        report_type=report_type
    )

    data_map = {
        entry.row_name: entry
        for entry in saved_entries
    }

    context = {
        "leading_indicator_rows": leading_indicator_rows,
        "lagging_indicator_rows": lagging_indicator_rows,
        "months": months,
        "selected_type": selected_type,
        "selected_plant": plant,
        "user_plants": user_plants,
        "current_year": year,
        "data_map": data_map
    }

    return render(
        request,
        "environmental_mis/safety_indicator.html",
        context
    )

# ===============================
# SAFETY REPORT SAVE 
# ===============================

def save_safety_indicator(request):
    if request.method != "POST":
        return JsonResponse({"status": "error"}, status=400)

    data = json.loads(request.body)

    plant_id = data.get("plant_id")
    report_type = data.get("report_type")
    year = timezone.now().year
    
    rows = data.get("rows", [])

    plant = Plant.objects.get(id=plant_id)

    if report_type == "LEADING_INDICATOR":
        report_type = "LEADING_INDICATOR"
    else:
        report_type = "LAGGING_INDICATOR"

    months = [
        "jan_qty","feb_qty","mar_qty","apr_qty","may_qty","jun_qty",
        "jul_qty","aug_qty","sep_qty","oct_qty","nov_qty","dec_qty"
    ]

    for row in rows:

        row_name = row.get("row_name")

        # Convert empty values to 0
        for m in months:
            val = row.get(m)
            row[m] = float(val) if val not in ["", None] else 0

        #calculate totals
        row["total_quantity"] = sum(row[m] for m in months)

        SafetyIndicatorEntry.objects.update_or_create(
            plant=plant,
            year=year,
            report_type=report_type,
            row_name=row_name,
            defaults=row
        )

    return JsonResponse({"status": "success"})


# ===============================
# ANALYTICAL DASHBOARD ENVIRONMENT
# ===============================

def environment_dashboard(request):

    user_plants = Plant.objects.all()

    plant_id = request.GET.get("plant_id")
    selected_type = request.GET.get("type", "MANUFACTURING")

    year = timezone.now().year

    report_type = (
        "MANUFACTURING_ENV"
        if selected_type == "MANUFACTURING"
        else "NON_MANUFACTURING_ENV"
    )

    # ================= PLANT FILTER =================

    if plant_id and plant_id != "all":

        selected_plant = Plant.objects.filter(id=plant_id).first()

        entries = EnvironmentEntry.objects.filter(
            plant=selected_plant,
            year=year,
            report_type=report_type
        )

    else:

        selected_plant = None

        entries = EnvironmentEntry.objects.filter(
            year=year,
            report_type=report_type
        )

    # ================= DATA AGGREGATION =================

    data = {}

    for e in entries:

        if e.row_name not in data:

            data[e.row_name] = e

        else:

            existing = data[e.row_name]

            months = [
                "jan_qty","feb_qty","mar_qty",
                "apr_qty","may_qty","jun_qty",
                "jul_qty","aug_qty","sep_qty",
                "oct_qty","nov_qty","dec_qty"
            ]

            for m in months:
                setattr(existing, m, getattr(existing, m) + getattr(e, m))

            existing.total_quantity += e.total_quantity

    months = [
        "Jan","Feb","Mar","Apr","May","Jun",
        "Jul","Aug","Sep","Oct","Nov","Dec"
    ]

    def get_monthly(row_key):

        row = data.get(row_key)

        if not row:
            return [0]*12

        return [
            row.jan_qty,row.feb_qty,row.mar_qty,
            row.apr_qty,row.may_qty,row.jun_qty,
            row.jul_qty,row.aug_qty,row.sep_qty,
            row.oct_qty,row.nov_qty,row.dec_qty
        ]

    # ================= KPI =================

    if selected_type == "MANUFACTURING":

        water = data.get("total_water_intake")
        energy = data.get("electric_power_volume")
        cost = data.get("total_ERA_bill_cost")

    else:

        water = data.get("Non_total_kl")
        energy = data.get("Non_total_energy_power_volume")
        cost = data.get("Non_total_energy_power_cost")

    # ================= PLANT ENERGY COMPARISON =================

    plant_labels = []
    plant_energy = []
    plant_water = []

    for plant in user_plants:

        if selected_type == "MANUFACTURING":

            energy_row = "electric_power_volume"
            water_row = "total_water_intake"

        else:

            energy_row = "Non_total_energy_power_volume"
            water_row = "Non_total_kl"


        energy_entry = EnvironmentEntry.objects.filter(
            plant=plant,
            year=year,
            report_type=report_type,
            row_name=energy_row
        ).first()

        water_entry = EnvironmentEntry.objects.filter(
            plant=plant,
            year=year,
            report_type=report_type,
            row_name=water_row
        ).first()


        plant_labels.append(plant.name)

        plant_energy.append(
            energy_entry.total_quantity if energy_entry else 0
        )

        plant_water.append(
            water_entry.total_quantity if water_entry else 0
        )

    # ================= ENERGY SOURCE =================

    energy_sources = [

        data.get("disel_mobile_volume").total_quantity if data.get("disel_mobile_volume") else 0,

        data.get("disel_stationary_volume").total_quantity if data.get("disel_stationary_volume") else 0,

        data.get("electric_power_volume").total_quantity if data.get("electric_power_volume") else 0,

        data.get("natural_gas_volume").total_quantity if data.get("natural_gas_volume") else 0,

        data.get("renewable_energy_solar_used_volume").total_quantity if data.get("renewable_energy_solar_used_volume") else 0
    ]

    if selected_type == "MANUFACTURING":

        water_cost = data.get("total_mnm_water_cost")

        cost_data = [
            water_cost.total_quantity if water_cost else 0,
            cost.total_quantity if cost else 0
        ]

    else:

        water_cost = data.get("Non_water_cost_rs")

        cost_data = [
            water_cost.total_quantity if water_cost else 0,
            cost.total_quantity if cost else 0
        ]

    def r2(val):
        return round(val or 0, 2)

    context = {

        "user_plants": user_plants,
        "selected_plant": selected_plant,
        "selected_type": selected_type,

        "months": months,

        "kpi_water": r2(water.total_quantity if water else 0),
        "kpi_energy": r2(energy.total_quantity if energy else 0),
        "kpi_cost": r2(cost.total_quantity if cost else 0),

        "water_monthly": get_monthly(
            "total_water_intake"
            if selected_type=="MANUFACTURING"
            else "Non_total_kl"
        ),

        "energy_monthly": get_monthly(
            "electric_power_volume"
            if selected_type=="MANUFACTURING"
            else "Non_total_energy_power_volume"
        ),

        "energy_sources": energy_sources,
        "cost_data": cost_data,

        "plant_labels": plant_labels,
        "plant_energy": plant_energy,
        "plant_water": plant_water,

        "inlet_cod": get_monthly("inlet_effluent_cod"),
        "treated_cod": get_monthly("treated_effuent_tank_cod"),

        "inlet_bod": get_monthly("inlet_effluent_bod"),
        "treated_bod": get_monthly("treated_effuent_tank_bod"),
    }

    return render(
        request,
        "environmental_mis/environment_dashboard.html",
        context
    )

# ===============================
# ANALYTICAL DASHBOARD SAFETY
# ===============================

def download_safety_excel(request):

    plant_id = request.GET.get("plant_id")
    selected_type = request.GET.get("type", "LEADING_INDICATOR")

    plant = Plant.objects.get(id=plant_id)
    year = timezone.now().year

    structure = (
        LEADING_INDICATOR
        if selected_type == "LEADING_INDICATOR"
        else LAGGING_INDICATOR
    )

    saved_entries = SafetyIndicatorEntry.objects.filter(
        plant=plant,
        year=year,
        report_type=selected_type
    )

    data_map = {e.row_name: e for e in saved_entries}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Safety Indicator"

    header_fill = PatternFill(start_color="79A1C9", end_color="79A1C9", fill_type="solid")
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    months = [
        "jan","feb","mar","apr","may","jun",
        "jul","aug","sep","oct","nov","dec"
    ]

    # ================= HEADERS =================

    headers = [
        "SL NO",
        "CATEGORY",
        "SUB CATEGORY",
        "QUESTION",
        "2025",
        "Jan","Feb","Mar","Apr","May","Jun",
        "Jul","Aug","Sep","Oct","Nov","Dec",
        "TOTAL"
    ]

    title = f"{plant.name} - {selected_type.replace('_',' ')} - {year}"

    last_col = get_column_letter(len(headers))

    ws.merge_cells(f"A1:{last_col}1")

    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center

    ws.row_dimensions[1].height = 28

    ws.append([])
    ws.append(headers)

    for c in ws[3]:
        c.font = bold_font
        c.fill = header_fill
        c.alignment = center

    # ================= DATA =================

    sl = 0
    last_section = None
    current_section = None

    for row in structure:

        if row.get("section"):
            current_section = row["section"]

        section = current_section

        if section != last_section:
            sl += 1
            sr = sl
            category = section
            last_section = section
        else:
            sr = ""
            category = ""

        entry = data_map.get(row["row_key"])

        month_values = []
        total = 0

        for m in months:
            val = getattr(entry, f"{m}_qty", 0) if entry else 0
            month_values.append(val)
            total += val

        excel_row = [

            sr,
            category,
            row.get("sub_section",""),
            row.get("question",""),

            getattr(entry,"year_2025",0) if entry else 0,

            *month_values,

            total
        ]

        ws.append(excel_row)

    # ================= COLUMN WIDTH =================

    ws.column_dimensions["A"].width = 6

    for i, column in enumerate(ws.columns, 1):

        if i == 1:
            continue

        max_length = 0
        letter = get_column_letter(i)

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[letter].width = max_length + 3

    # ================= DOWNLOAD =================

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    filename = f"safety_indicator_{plant.name}_{year}.xlsx"

    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)

    return response

# ===============================
# ANALYTICAL DASHBOARD WASTE
# ===============================

def waste_dashboard(request):

    user_plants = Plant.objects.all()

    plant_id = request.GET.get("plant_id")
    waste_type = request.GET.get("waste_type", "MANUFACTURING")

    year = timezone.now().year


    # safer mapping
    report_map = {
        "MANUFACTURING": ["MANUFACTURING", "MANUFACTURING_WASTE"],
        "NON_MANUFACTURING": ["NON_MANUFACTURING", "NON_MANUFACTURING_WASTE"],
    }

    report_values = report_map.get(waste_type, ["MANUFACTURING"])


    if plant_id:
        selected_plant = Plant.objects.filter(id=plant_id).first()
    else:
        selected_plant = user_plants.first()


    summaries = WasteSummary.objects.filter(
        plant=selected_plant,
        report_type__in=report_values
    )


    summary_map = {s.summary_type: s for s in summaries}


    months = [
        "jan","feb","mar",
        "apr","may","jun",
        "jul","aug","sep",
        "oct","nov","dec"
    ]


    def monthly(summary_type):

        s = summary_map.get(summary_type)

        if not s:
            return [0]*12

        return [
            float(getattr(s, f"{m}_qty") or 0)
            for m in months
        ]


    def total(summary_type):

        s = summary_map.get(summary_type)

        if not s:
            return 0

        return float(s.total_quantity or 0)


    context = {

        "user_plants": user_plants,
        "selected_plant": selected_plant,
        "selected_type": waste_type,
        "year": year,

        "total_waste": total("GRAND_WASTE"),
        "nonhaz_total": total("NON_HAZ"),
        "haz_total": total("HAZ_PROCESS"),
        "ewaste_total": total("E_WASTE"),
        "diversion_total": total("DIVERSION_RATE"),

        "nonhaz_unit": total("NON_HAZ_UNIT"),
        "haz_unit": total("PROC_HAZ_UNIT"),

        "grand_data": monthly("GRAND_WASTE"),
        "nonhaz_data": monthly("NON_HAZ"),
        "haz_data": monthly("HAZ_PROCESS"),
        "ewaste_data": monthly("E_WASTE"),
        "diversion_data": monthly("DIVERSION_RATE"),
    }


    return render(
        request,
        "environmental_mis/waste_dashboard.html",
        context
    )


# ===============================
# WASTE EXCEL
# ===============================

def download_waste_excel(request):

    plant_id = request.GET.get("plant_id")
    waste_type = request.GET.get("type", "MANUFACTURING")

    plant = Plant.objects.get(id=plant_id)
    year = timezone.now().year

    # ================= MAIN DATA =================

    report_data = WasteReportData.objects.filter(
        plant=plant,
        year=year,
        report_type=waste_type
    )

    data_map = {d.row_name: d for d in report_data}

    # ================= SUMMARY DATA =================

    summary_data = WasteSummary.objects.filter(
        plant=plant,
        year=year,
        report_type=waste_type
    )

    summary_map = {s.summary_type: s for s in summary_data}

    # ================= WORKBOOK =================

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Waste Report"

    # ================= STYLES =================

    header_fill = PatternFill(start_color="79A1C9", end_color="79A1C9", fill_type="solid")
    total_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    # ================= HEADERS =================

    headers = [
        "SR.NO","SECTION","SCRAP","PART CODE","UNIT","TREATMENT",

        "Jan QTY","Jan COST",
        "Feb QTY","Feb COST",
        "Mar QTY","Mar COST",
        "I Q (QTY)","I Q (COST)",

        "Apr QTY","Apr COST",
        "May QTY","May COST",
        "Jun QTY","Jun COST",
        "II Q (QTY)","II Q (COST)",

        "Jul QTY","Jul COST",
        "Aug QTY","Aug COST",
        "Sep QTY","Sep COST",
        "III Q (QTY)","III Q (COST)",

        "Oct QTY","Oct COST",
        "Nov QTY","Nov COST",
        "Dec QTY","Dec COST",
        "IV Q (QTY)","IV Q (COST)",

        "TOTAL QTY","TOTAL COST"
    ]

    # ================= TITLE =================

    title = f"{plant.name} - MONTHLY - {waste_type.upper()} WASTE DATA - {year} Year"

    last_col = get_column_letter(len(headers))

    ws.merge_cells(f"A1:{last_col}1")

    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center_align

    title_cell.fill = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid"
    )

    ws.row_dimensions[1].height = 28

    # ================= HEADER ROW =================

    ws.append([])
    ws.append(headers)

    for cell in ws[3]:
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center_align

    # ================= ROW DATA =================

    sl = 0
    last_section = None
    current_section = None

    for row in MANUFACTURING_WASTE_STRUCTURE:

        saved = data_map.get(row["row_key"])

        # if section is empty, reuse previous section
        if row["section"]:
            current_section = str(row["section"]).strip()

        section = current_section

        # increase SR only when section changes
        if section != last_section:
            sl += 1
            sr_no = sl
            last_section = section
        else:
            sr_no = ""

        jan = saved.jan_qty if saved else 0
        feb = saved.feb_qty if saved else 0
        mar = saved.mar_qty if saved else 0

        apr = saved.apr_qty if saved else 0
        may = saved.may_qty if saved else 0
        jun = saved.jun_qty if saved else 0

        jul = saved.jul_qty if saved else 0
        aug = saved.aug_qty if saved else 0
        sep = saved.sep_qty if saved else 0

        octv = saved.oct_qty if saved else 0
        nov = saved.nov_qty if saved else 0
        dec = saved.dec_qty if saved else 0

        q1 = jan + feb + mar
        q2 = apr + may + jun
        q3 = jul + aug + sep
        q4 = octv + nov + dec

        total_qty = q1 + q2 + q3 + q4

        excel_row = [

            sr_no,
            section,
            row["question"],
            saved.part_code if saved else "",
            row["unit"],
            row["treatment_type"],

            jan, saved.jan_cost if saved else 0,
            feb, saved.feb_cost if saved else 0,
            mar, saved.mar_cost if saved else 0,

            q1, 0,

            apr, saved.apr_cost if saved else 0,
            may, saved.may_cost if saved else 0,
            jun, saved.jun_cost if saved else 0,

            q2, 0,

            jul, saved.jul_cost if saved else 0,
            aug, saved.aug_cost if saved else 0,
            sep, saved.sep_cost if saved else 0,

            q3, 0,

            octv, saved.oct_cost if saved else 0,
            nov, saved.nov_cost if saved else 0,
            dec, saved.dec_cost if saved else 0,

            q4, 0,

            total_qty,
            saved.total_cost if saved else 0
        ]

        ws.append(excel_row)

        if row["question"] == "TOTAL QUANTITY":
            for cell in ws[ws.max_row]:
                cell.fill = total_fill
                cell.font = bold_font

    # ================= COLUMN WIDTH =================

    ws.column_dimensions['A'].width = 5

    for i, column in enumerate(ws.columns, 1):

        if i == 1:
            continue

        max_length = 0
        column_letter = get_column_letter(i)

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = max_length + 3

    # ================= SUMMARY SHEET =================

    summary_ws = wb.create_sheet(title="Waste Summary")

    summary_headers = [
        "Waste Type",

        "Jan QTY","Jan COST",
        "Feb QTY","Feb COST",
        "Mar QTY","Mar COST",
        "Q1 QTY","Q1 COST",

        "Apr QTY","Apr COST",
        "May QTY","May COST",
        "Jun QTY","Jun COST",
        "Q2 QTY","Q2 COST",

        "Jul QTY","Jul COST",
        "Aug QTY","Aug COST",
        "Sep QTY","Sep COST",
        "Q3 QTY","Q3 COST",

        "Oct QTY","Oct COST",
        "Nov QTY","Nov COST",
        "Dec QTY","Dec COST",
        "Q4 QTY","Q4 COST",

        "TOTAL QTY","TOTAL COST"
    ]

    summary_ws.append(summary_headers)

    for cell in summary_ws[1]:
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center_align

    summary_types = [
        "NON_HAZ",
        "HAZ_PROCESS",
        "E_WASTE",
        "GRAND_WASTE",
        "PRODUCTION",
        "NON_HAZ_UNIT",
        "PROC_HAZ_UNIT",
        "NON_PROC_HAZ",
        "DIVERSION_RATE"
    ]

    months = [
        "jan","feb","mar",
        "apr","may","jun",
        "jul","aug","sep",
        "oct","nov","dec"
    ]

    for s_type in summary_types:

        s = summary_map.get(s_type)

        row_data = [s_type]

        for m in months:
            row_data.append(getattr(s, f"{m}_qty", 0) if s else 0)
            row_data.append(getattr(s, f"{m}_cost", 0) if s else 0)

        row_data.append(getattr(s,"q1_quantity",0) if s else 0)
        row_data.append(getattr(s,"q1_cost",0) if s else 0)

        row_data.append(getattr(s,"q2_quantity",0) if s else 0)
        row_data.append(getattr(s,"q2_cost",0) if s else 0)

        row_data.append(getattr(s,"q3_quantity",0) if s else 0)
        row_data.append(getattr(s,"q3_cost",0) if s else 0)

        row_data.append(getattr(s,"q4_quantity",0) if s else 0)
        row_data.append(getattr(s,"q4_cost",0) if s else 0)

        row_data.append(getattr(s,"total_quantity",0) if s else 0)
        row_data.append(getattr(s,"total_cost",0) if s else 0)

        summary_ws.append(row_data)

    for i, column in enumerate(summary_ws.columns, 1):

        max_length = 0
        column_letter = get_column_letter(i)

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        summary_ws.column_dimensions[column_letter].width = max_length + 3

    # ================= DOWNLOAD =================

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    filename = f"waste_report_{plant.name}_{year}.xlsx"

    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)

    return response

# ===============================
# ENVIRONMENT EXCEL
# ===============================

def download_environment_excel(request):

    plant_id = request.GET.get("plant_id")
    selected_type = request.GET.get("type", "MANUFACTURING")

    plant = Plant.objects.get(id=plant_id)
    year = timezone.now().year

    report_type = (
        "MANUFACTURING_ENV"
        if selected_type == "MANUFACTURING"
        else "NON_MANUFACTURING_ENV"
    )

    saved_entries = EnvironmentEntry.objects.filter(
        plant=plant,
        year=year,
        report_type=report_type
    )

    data_map = {e.row_name: e for e in saved_entries}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Environment Report"

    header_fill = PatternFill(start_color="79A1C9", end_color="79A1C9", fill_type="solid")
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    months = [
        "jan","feb","mar","apr","may","jun",
        "jul","aug","sep","oct","nov","dec"
    ]

    # ================= HEADERS =================

    headers = [
        "SL NO",
        "CATEGORY",
        "SUB CATEGORY",
        "QUESTION",
        "UNIT CATEGORY",
        "UNIT",
        "2024",
        "2025",
        "Jan","Feb","Mar","Apr","May","Jun",
        "Jul","Aug","Sep","Oct","Nov","Dec",
        "TOTAL"
    ]

    title = f"{plant.name} - {selected_type} ENVIRONMENT REPORT - {year}"

    last_col = get_column_letter(len(headers))
    ws.merge_cells(f"A1:{last_col}1")

    cell = ws["A1"]
    cell.value = title
    cell.font = Font(bold=True, size=14)
    cell.alignment = center

    ws.append([])
    ws.append(headers)

    for c in ws[3]:
        c.font = bold_font
        c.fill = header_fill
        c.alignment = center

    # ================= STRUCTURE =================

    structure = (
        MANUFACTURING_ENVIRONMENT_STRUCTURE
        if selected_type == "MANUFACTURING"
        else NON_MANUFACTURING_ENVIRONMENT_STRUCTURE
    )

    sl = 0
    last_section = None
    current_section = None

    for row in structure:

        if row["section"]:
            current_section = row["section"]

        section = current_section

        if section != last_section:
            sl += 1
            sr = sl
            last_section = section
        else:
            sr = ""

        entry = data_map.get(row["row_key"])

        month_values = []
        total = 0

        for m in months:
            val = getattr(entry, f"{m}_qty", 0) if entry else 0
            month_values.append(val)
            total += val

        excel_row = [

            sr,
            section,
            row.get("sub_section",""),
            row.get("question",""),
            row.get("unit_category",""),
            row.get("unit",""),

            getattr(entry,"year_2024",0) if entry else 0,
            getattr(entry,"year_2025",0) if entry else 0,

            *month_values,

            total
        ]

        ws.append(excel_row)

    # ================= WIDTH =================

    ws.column_dimensions["A"].width = 6

    for i, column in enumerate(ws.columns, 1):

        if i == 1:
            continue

        max_len = 0
        letter = get_column_letter(i)

        for cell in column:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[letter].width = max_len + 3

    # ================= DOWNLOAD =================

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    filename = f"environment_report_{plant.name}_{year}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)

    return response

# ===============================
# SAFETY EXCEL
# ===============================

def safety_dashboard(request):

    user_plants = Plant.objects.all()

    plant_id = request.GET.get("plant_id")
    selected_type = request.GET.get("type", "LEADING_INDICATOR")

    year = timezone.now().year

    if plant_id:
        selected_plant = Plant.objects.filter(id=plant_id).first()
    else:
        selected_plant = user_plants.first()

    entries = SafetyIndicatorEntry.objects.filter(
        plant=selected_plant,
        year=year,
        report_type=selected_type
    )

    data_map = {e.row_name: e for e in entries}

    months = [
        "jan","feb","mar",
        "apr","may","jun",
        "jul","aug","sep",
        "oct","nov","dec"
    ]

    def monthly(row_key):
        entry = data_map.get(row_key)
        if not entry:
            return [0]*12

        return [
            float(getattr(entry, f"{m}_qty") or 0)
            for m in months
        ]

    def total(row_key):
        entry = data_map.get(row_key)
        if not entry:
            return 0
        return float(entry.total_quantity or 0)


    # ----------------------
    # Leading calculations
    # ----------------------

    walk_plan_total = total("management_safety_walk_plan")
    walk_actual_total = total("management_safety_walk_actual")

    walk_compliance = 0
    if walk_plan_total > 0:
        walk_compliance = round((walk_actual_total / walk_plan_total) * 100, 1)


    # ----------------------
    # Context
    # ----------------------

    context = {

        "user_plants": user_plants,
        "selected_plant": selected_plant,
        "selected_type": selected_type,
        "year": year,

        # KPI
        "lti_total": total("lost_time_incident"),
        "fatality_total": total("fatality"),
        "hipo_total": total("hipo_events_contractor_employees"),
        "days_lost_total": total("no_of_days_lost"),

        "walk_plan_total": walk_plan_total,
        "walk_actual_total": walk_actual_total,
        "audit_actual_total": total("safety_audit_conducted"),
        "meeting_compliance_avg": walk_compliance,


        # ---------- Lagging Charts ----------

        "lti_data": monthly("lost_time_incident"),
        "recordable_data": monthly("recordable_cases_only"),
        "first_aid_data": monthly("first_aid_cases_fac"),

        "fatality_data": monthly("fatality"),
        "hipo_data": monthly("hipo_events_contractor_employees"),
        "days_lost_data": monthly("no_of_days_lost"),


        # ---------- Leading Charts ----------

        "walk_plan": monthly("management_safety_walk_plan"),
        "walk_actual": monthly("management_safety_walk_actual"),

        "audit_plan": monthly("safety_audit_planned"),
        "audit_actual": monthly("safety_audit_conducted"),

        "meeting_compliance": monthly("level_1_meeting_compliance"),

        "observations_data": monthly("management_safety_walk_observations"),

        "committee_plan": monthly("safety_committee_meetings_planned"),
        "committee_actual": monthly("safety_committee_meetings_actual"),

        "training_data": monthly("temp_employee_3days_training"),
    }

    return render(
        request,
        "environmental_mis/safety_dashboard.html",
        context
    ) 