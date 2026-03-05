from multiprocessing import context
from urllib.parse import urlencode
from django.contrib.auth.mixins import LoginRequiredMixin,UserPassesTestMixin
from django.views.generic import FormView, ListView, CreateView, UpdateView, DetailView, TemplateView,DeleteView
from django.urls import reverse, reverse_lazy
from django.shortcuts import get_object_or_404, redirect
from django.db.models import Q, Count, Sum
from django.http import JsonResponse
from apps.organizations.models import *
from .models import *
from .forms import *
from .utils import generate_incident_pdf
from django.http import HttpResponse
from django.db.models.functions import TruncMonth
from django.views.generic import UpdateView, TemplateView
from django.contrib import messages
from django.utils import timezone
from apps.accidents.models import IncidentType
from apps.notifications import *
import datetime
from django.db.models import Q
import json
import openpyxl
from django.shortcuts import render
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from django.conf import settings  
from django.conf.urls.static import static  
from apps.common.image_utils import compress_image

from .forms import IncidentAttachmentForm # <-- Import the new form
from django.views.generic import UpdateView
from django.views.generic import ListView
from .models import IncidentActionItem
from django.db.models import Exists, OuterRef



class IncidentTypeListView(LoginRequiredMixin, ListView):
    """List all incident types with search functionality"""
    model = IncidentType
    template_name = 'accidents/incident_type_list.html'
    context_object_name = 'incident_types'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = IncidentType.objects.annotate(
            incident_count=Count('incidents')
        ).order_by('name')
        
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(code__icontains=search_query) |
                Q(description__icontains=search_query)
            )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        total_injuries = self.object_list.aggregate(total=Sum('incident_count')) ['total'] or 0
        context['total_injuries'] = total_injuries 
        context['search_query'] = self.request.GET.get('search', '')
        return context


class IncidentTypeCreateView(LoginRequiredMixin, CreateView):
    """Create a new incident type"""
    model = IncidentType
    form_class = IncidentTypeForm
    template_name = 'accidents/incident_type_form.html'
    success_url = reverse_lazy('accidents:incident_type_list')
    
    def form_valid(self, form):
        incident_type = form.save(commit=False)
        incident_type.created_by = self.request.user
        incident_type.save()
        messages.success(
            self.request, 
            f'Incident Type "{incident_type.name}" created successfully!'
        )
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['action'] = 'Create'
        return context


class IncidentTypeUpdateView(LoginRequiredMixin, UpdateView):
    """Update an existing incident type"""
    model = IncidentType
    form_class = IncidentTypeForm
    template_name = 'accidents/incident_type_form.html'
    success_url = reverse_lazy('accidents:incident_type_list')
    
    def form_valid(self, form):
        messages.success(
            self.request,
            f'Incident Type "{self.object.name}" updated successfully!'
        )
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['action'] = 'Update'
        context['incident_type'] = self.object
        return context


class IncidentTypeDeleteView(LoginRequiredMixin, DeleteView):
    """Delete an incident type"""
    model = IncidentType
    template_name = 'accidents/incident_type_confirm_delete.html'
    success_url = reverse_lazy('accidents:incident_type_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['incident_count'] = self.object.incidents.count()
        return context
    
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        incident_count = self.object.incidents.count()
        
        if incident_count > 0:
            messages.error(
                request,
                f'Cannot delete "{self.object.name}". It is being used by {incident_count} incident(s).'
            )
            return redirect('accidents:incident_type_list')
        
        incident_type_name = self.object.name
        success_url = self.get_success_url()
        self.object.delete()
        
        messages.success(
            request,
            f'Incident Type "{incident_type_name}" deleted successfully!'
        )
        return redirect(success_url)
class IncidentDashboardView(LoginRequiredMixin, TemplateView):
    """Incident Management Dashboard"""
    template_name = 'accidents/dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get incidents based on user role
        if self.request.user.is_superuser or self.request.user.role.name == 'ADMIN':
            incidents = Incident.objects.all()
        elif self.request.user.plant:
            incidents = Incident.objects.filter(plant=self.request.user.plant)
        else:
            incidents = Incident.objects.filter(reported_by=self.request.user)
        
        # Statistics
        context['total_incidents'] = incidents.count()
        context['open_incidents'] = incidents.exclude(status='CLOSED').count()
        context['this_month_incidents'] = incidents.filter(
            incident_date__month=datetime.date.today().month,
            incident_date__year=datetime.date.today().year
        ).count()
        context['investigation_pending'] = incidents.filter(
            investigation_required=True,
            investigation_completed_date__isnull=True
        ).count()
        
        # ✅ UPDATED: By Type using ForeignKey relationship
        # Get incident type IDs for common types
        try:
            lti_type = IncidentType.objects.get(code='LTI')
            context['lti_count'] = incidents.filter(incident_type__code='LTI').count()
        except IncidentType.DoesNotExist:
            context['lti_count'] = 0
        
        try:
            mtc_type = IncidentType.objects.get(code='MTC')
            context['mtc_count'] = incidents.filter(incident_type__code='MTC').count()
        except IncidentType.DoesNotExist:
            context['mtc_count'] = 0
        
        try:
            fa_type = IncidentType.objects.get(code='FA')
            context['fa_count'] = incidents.filter(incident_type__code='FA').count()
        except IncidentType.DoesNotExist:
            context['fa_count'] = 0
        
        try:
            hlfi_type = IncidentType.objects.get(code='HLFI')
            context['hlfi_count'] = incidents.filter(incident_type__code='HLFI').count()
        except IncidentType.DoesNotExist:
            context['hlfi_count'] = 0
        
        # Recent incidents
        context['recent_incidents'] = incidents.order_by('-reported_date')[:10]
        
        # Overdue investigations
        context['overdue_investigations'] = incidents.filter(
            investigation_required=True,
            investigation_completed_date__isnull=True,
            investigation_deadline__lt=datetime.date.today()
        )
        
        return context


class IncidentListView(LoginRequiredMixin, ListView):
    """List all incidents"""
    model = Incident
    template_name = 'accidents/incident_list.html'
    context_object_name = 'incidents'
    paginate_by = 20
    
    def get_queryset(self):
        # Get the current logged-in user
        user = self.request.user
        
        # Start with the base queryset, fetching related objects to optimize queries
        queryset = Incident.objects.select_related('plant', 'location', 'reported_by','incident_type').order_by('-incident_date', '-incident_time')
        
        # --- ROLE-BASED DATA FILTERING ---
        # Check if the user is a superuser or has an ADMIN role
        if user.is_superuser or (hasattr(user, 'role') and user.role and user.role.name == 'ADMIN'):
            # No filtering needed; they can see all records
            pass
        # Check if the user has an EMPLOYEE role
        elif hasattr(user, 'role') and user.role and user.role.name == 'EMPLOYEE':
            # Filter the queryset to show only records reported by the current user
            queryset = queryset.filter(reported_by=user)
        # Check if the user is associated with a specific plant (for roles like PLANT HEAD, etc.)
        elif user.plant:
            queryset = queryset.filter(plant=user.plant)
        else:
            # As a fallback, if no specific role logic applies, show only self-reported records
            queryset = queryset.filter(reported_by=user)
        
        # --- SEARCH AND FILTER LOGIC ---
        # This part remains the same and applies on top of the role-filtered queryset
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(report_number__icontains=search) |
                Q(affected_person_name__icontains=search)
            )
        
        incident_type = self.request.GET.get('incident_type')
        if incident_type:
            queryset = queryset.filter(incident_type_id=incident_type)
        
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        plant = self.request.GET.get('plant')
        if plant:
            queryset = queryset.filter(plant_id=plant)
        
        date_from = self.request.GET.get('date_from')
        if date_from:
            queryset = queryset.filter(incident_date__gte=date_from)
            
        date_to = self.request.GET.get('date_to')
        if date_to:
            queryset = queryset.filter(incident_date__lte=date_to)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from apps.organizations.models import Plant
        context['plants'] = Plant.objects.filter(is_active=True)
        context['incident_types'] = IncidentType.objects.filter(is_active=True).order_by('name')
        context['status_choices'] = Incident.STATUS_CHOICES
        context['search_query'] = self.request.GET.get('search', '')
        context['selected_incident_type'] = self.request.GET.get('incident_type', '')
        context['selected_status'] = self.request.GET.get('status', '')
        context['selected_plant'] = self.request.GET.get('plant', '')
        return context


# Updated IncidentCreateView to handle unsafe acts and conditions
class IncidentCreateView(LoginRequiredMixin, CreateView):
    model = Incident
    form_class = IncidentReportForm
    template_name = 'accidents/incident_create.html'
    success_url = reverse_lazy('accidents:incident_list')

    def get_form_kwargs(self):
        """
        Passes the current request's user to the form's __init__ method.
        This is CRUCIAL for the form logic to work.
        """
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        """
        Adds the user's location assignments to the template context.
        This allows the template to conditionally render fields as readonly or dropdowns.
        """
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        # Pass QuerySets of assigned locations to the template
        context['user_assigned_plants'] = user.assigned_plants.filter(is_active=True)
        
        # Pass other assignments for the template logic to use
        # The template can now check the count of each of these
        if context['user_assigned_plants'].count() == 1:
            plant = context['user_assigned_plants'].first()
            context['user_assigned_zones'] = user.assigned_zones.filter(is_active=True, plant=plant)
            if context['user_assigned_zones'].count() == 1:
                zone = context['user_assigned_zones'].first()
                context['user_assigned_locations'] = user.assigned_locations.filter(is_active=True, zone=zone)
                if context['user_assigned_locations'].count() == 1:
                    location = context['user_assigned_locations'].first()
                    context['user_assigned_sublocations'] = user.assigned_sublocations.filter(is_active=True, location=location)
                else:
                    context['user_assigned_sublocations'] = user.assigned_sublocations.none() # Or all if needed
            else:
                context['user_assigned_locations'] = user.assigned_locations.none()
                context['user_assigned_sublocations'] = user.assigned_sublocations.none()
        else:
            context['user_assigned_zones'] = user.assigned_zones.none()
            context['user_assigned_locations'] = user.assigned_locations.none()
            context['user_assigned_sublocations'] = user.assigned_sublocations.none()

        context['active_incident_types'] = IncidentType.objects.filter(is_active=True)
        context['departments'] = Department.objects.filter(is_active=True).order_by('name')
        context['cancel_url'] = (self.request.GET.get('next') or self.request.META.get('HTTP_REFERER') or '/')
        return context
    
    def form_valid(self, form):
        """
        Process the valid form, set the reporter, and handle location data.
        """
        incident = form.save(commit=False)
        incident.reported_by = self.request.user
        
        user = self.request.user

        # Manually set location fields if they are single-assigned and might not be in the form post data
        # (e.g., if we use readonly fields instead of disabled dropdowns).
        if user.assigned_plants.count() == 1 and not form.cleaned_data.get('plant'):
            incident.plant = user.assigned_plants.first()
        
        if user.assigned_zones.count() == 1 and not form.cleaned_data.get('zone'):
            incident.zone = user.assigned_zones.first()

        if user.assigned_locations.count() == 1 and not form.cleaned_data.get('location'):
            incident.location = user.assigned_locations.first()

        if user.assigned_sublocations.count() == 1 and not form.cleaned_data.get('sublocation'):
            incident.sublocation = user.assigned_sublocations.first()

        # Handle JSON fields from hidden inputs
        incident.affected_body_parts = json.loads(self.request.POST.get('affected_body_parts_json', '[]'))
        incident.unsafe_acts = json.loads(self.request.POST.get('unsafe_acts_json', '[]'))
        incident.unsafe_conditions = json.loads(self.request.POST.get('unsafe_conditions_json', '[]'))
        incident.unsafe_acts_other = self.request.POST.get('unsafe_acts_other', '').strip()
        incident.unsafe_conditions_other = self.request.POST.get('unsafe_conditions_other', '').strip()
        
        incident.save()
        self.object = incident
        form.save_m2m()

        # Handle photo uploads
        photos = self.request.FILES.getlist('photos')
        for photo in photos:
            compressed_photo = compress_image(photo)
            IncidentPhoto.objects.create(
                incident=incident,
                photo=compressed_photo,
                photo_type='INCIDENT_SCENE',
                uploaded_by=self.request.user
            )
        
        # ===== ADD NOTIFICATION HERE - AFTER INCIDENT IS SAVED =====
        # print("\n\n" + "#" * 70)
        # print("VIEW: INCIDENT SAVED SUCCESSFULLY")
        # print("#" * 70)
        # print(f"Report Number: {self.object.report_number}")
        # print(f"Incident ID: {self.object.id}")
        # print(f"Plant: {self.object.plant}")
        # print(f"Location: {self.object.location}")
        # print("#" * 70)
        # print("\nVIEW: Calling notify_incident_reported()...")

        try:
            # ✅ NEW: Use NotificationService instead of old notification_utils
            from apps.notifications.services import NotificationService
            
            NotificationService.notify(
                content_object=self.object,
                notification_type='INCIDENT_REPORTED',
                module='INCIDENT'
            )
            
            print("\nVIEW: ✅ Notifications sent successfully")
        except Exception as e:
            print(f"\nVIEW: ❌ ERROR in notification system: {e}")
            import traceback
            traceback.print_exc()
    
        print("\n" + "#" * 70 + "\n\n")
        
        messages.success(
            self.request,
            f'Incident {self.object.report_number} reported successfully! Investigation required within 7 days.'
        )
        
        return redirect(self.get_success_url())

    def form_invalid(self, form):
        messages.error(self.request, 'Please correct the errors below.')
        return super().form_invalid(form)

class IncidentDetailView(LoginRequiredMixin, DetailView):
    """View incident details"""
    model = Incident
    template_name = 'accidents/incident_detail.html'
    context_object_name = 'incident'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['photos'] = self.object.photos.all()
        context['action_items'] = self.object.action_items.all()
        context['cancel_url'] = (self.request.GET.get('next') or self.request.META.get('HTTP_REFERER') or '/')
        
        try:
            context['investigation_report'] = self.object.investigation_report
        except:
            context['investigation_report'] = None
        
        return context


# ============================================================================
# REPLACE YOUR EXISTING IncidentUpdateView WITH THIS UPDATED VERSION
# ============================================================================

def get_zones_by_plant(request, plant_id):
    """
    Fetch zones for a given plant ID.
    Returns a JSON response for AJAX calls.
    """
    if not request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'error': 'Invalid request'}, status=400)
        
    try:
        zones = Zone.objects.filter(plant_id=plant_id, is_active=True).values('id', 'name', 'code')
        return JsonResponse({'zones': list(zones)})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_locations_by_zone(request, zone_id):
    """
    Fetch locations for a given zone ID.
    Returns a JSON response for AJAX calls.
    """
    if not request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'error': 'Invalid request'}, status=400)

    try:
        locations = Location.objects.filter(zone_id=zone_id, is_active=True).values('id', 'name', 'code')
        return JsonResponse({'locations': list(locations)})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_sublocations_by_location(request, location_id):
    """
    Fetch sub-locations for a given location ID.
    Returns a JSON response for AJAX calls.
    """
    if not request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'error': 'Invalid request'}, status=400)
        
    try:
        sublocations = SubLocation.objects.filter(location_id=location_id, is_active=True).values('id', 'name', 'code')
        return JsonResponse({'sublocations': list(sublocations)})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
class IncidentUpdateView(LoginRequiredMixin, UpdateView):
    """Update incident report"""
    model = Incident
    form_class = IncidentUpdateForm
    template_name = 'accidents/incident_update.html'
    
    def get_success_url(self):
        return reverse_lazy('accidents:incident_detail', kwargs={'pk': self.object.pk})

    def get_form_kwargs(self):
        """
        Passes the current request's user to the form's __init__ method.
        This is CRUCIAL for the form logic to work correctly.
        """
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def get_context_data(self, **kwargs):
        """
        Adds the user's location assignments and JSON data to the template context.
        """
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        # Pass QuerySets of assigned locations to the template
        context['user_assigned_plants'] = user.assigned_plants.filter(is_active=True)

        # This logic helps the template decide if there's only one option to show
        if context['user_assigned_plants'].count() == 1:
            plant = context['user_assigned_plants'].first()
            context['user_assigned_zones'] = user.assigned_zones.filter(is_active=True, plant=plant)
            if context['user_assigned_zones'].count() == 1:
                zone = context['user_assigned_zones'].first()
                context['user_assigned_locations'] = user.assigned_locations.filter(is_active=True, zone=zone)
                if context['user_assigned_locations'].count() == 1:
                    location = context['user_assigned_locations'].first()
                    context['user_assigned_sublocations'] = user.assigned_sublocations.filter(is_active=True, location=location)
                else:
                    context['user_assigned_sublocations'] = user.assigned_sublocations.none()
            else:
                context['user_assigned_locations'] = user.assigned_locations.none()
                context['user_assigned_sublocations'] = user.assigned_sublocations.none()
        else:
            context['user_assigned_zones'] = user.assigned_zones.none()
            context['user_assigned_locations'] = user.assigned_locations.none()
            context['user_assigned_sublocations'] = user.assigned_sublocations.none()
        
        # Add departments for the affected person dropdown
        context['departments'] = Department.objects.filter(is_active=True).order_by('name')
        
        context['active_incident_types'] = IncidentType.objects.filter(is_active=True)

        # ✅ START: ADDED CODE
        # Add JSON-stringified data for safe JS initialization in the template
        context['affected_body_parts_json'] = json.dumps(
            self.object.affected_body_parts or []
        )
        # ✅ END: ADDED CODE
        context['cancel_url'] = (self.request.GET.get('next') or self.request.META.get('HTTP_REFERER') or '/')
        
        return context
    
        
    #Incident Update Form
    def get_form(self, form_class=None):
        form = super().get_form(form_class)

        incident = self.object  
        form.initial['affected_body_parts_json'] = json.dumps(
            incident.affected_body_parts or []
        )

        form.initial['unsafe_acts_json'] = json.dumps(
            incident.unsafe_acts or []
        )

        form.initial['unsafe_conditions_json'] = json.dumps(
            incident.unsafe_conditions or []
        )
        form.initial['unsafe_acts_other'] = incident.unsafe_acts_other or ''
        form.initial['unsafe_conditions_other'] = incident.unsafe_conditions_other or ''
        if incident.affected_person:
            form.initial['affected_person_id'] = incident.affected_person.id

        return form

    
    def form_valid(self, form):
        # Handle affected person selection
        incident_type = form.cleaned_data.get('incident_type')
        if incident_type:
            form.instance.incident_type = incident_type

        affected_person_id = self.request.POST.get('affected_person_id', '').strip()
        if affected_person_id:
            try:
                affected_employee = User.objects.select_related('department').get(
                    id=affected_person_id,
                    is_active=True
                )
                form.instance.affected_person = affected_employee
                form.instance.affected_person_name = affected_employee.get_full_name()
                form.instance.affected_person_employee_id = affected_employee.employee_id or ''
                form.instance.affected_person_department = affected_employee.department
            except (User.DoesNotExist, ValueError):
                pass
        
        # Handle affected body parts JSON
        affected_body_parts_json = self.request.POST.get('affected_body_parts_json', '[]')
        try:
            form.instance.affected_body_parts = json.loads(affected_body_parts_json)
        except:
            form.instance.affected_body_parts = []
        
        # Handle unsafe acts JSON
        unsafe_acts_json = self.request.POST.get('unsafe_acts_json', '[]')
        try:
            form.instance.unsafe_acts = json.loads(unsafe_acts_json)
        except:
            form.instance.unsafe_acts = []
        
        # Handle unsafe acts other explanation
        form.instance.unsafe_acts_other = self.request.POST.get('unsafe_acts_other', '').strip()
        
        # Handle unsafe conditions JSON
        unsafe_conditions_json = self.request.POST.get('unsafe_conditions_json', '[]')
        try:
            form.instance.unsafe_conditions = json.loads(unsafe_conditions_json)
        except:
            form.instance.unsafe_conditions = []
        
        # Handle unsafe conditions other explanation
        form.instance.unsafe_conditions_other = self.request.POST.get('unsafe_conditions_other', '').strip()
        
        # Handle photo uploads
        photos = self.request.FILES.getlist('photos')
        for photo in photos:
            compressed_photo = compress_image(photo)
            IncidentPhoto.objects.create(
                incident=self.object,
                photo=compressed_photo,
                photo_type='INCIDENT_SCENE',
                uploaded_by=self.request.user
            )
        
        messages.success(self.request, f'Incident {self.object.report_number} updated successfully!')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Please correct the errors below.')
        print("Form errors:", form.errors)
        return super().form_invalid(form)

import json
from django.forms import inlineformset_factory
from django.contrib.auth import get_user_model


User = get_user_model()
class InvestigationReportCreateView(LoginRequiredMixin, CreateView):
    """
    Create investigation report and its associated action items.
    Handles 'Self Assign' and 'Forward to Others' logic for action items.
    """
    model = IncidentInvestigationReport
    form_class = IncidentInvestigationReportForm
    template_name = 'accidents/investigation_report_create.html'

    def dispatch(self, request, *args, **kwargs):
        self.incident = get_object_or_404(Incident, pk=self.kwargs['incident_pk'])
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['incident'] = self.incident
        ActionItemFormSet = inlineformset_factory(
            Incident,
            IncidentActionItem,
            form=IncidentActionItemForm,
            extra=1,
            can_delete=True # Set to True to handle removals
        )
        if self.request.POST:
            context['action_item_formset'] = ActionItemFormSet(
                self.request.POST, self.request.FILES, # Add self.request.FILES for attachments
                instance=self.incident,
                prefix='actionitems',
                form_kwargs={'incident': self.incident}
            )
        else:
            context['action_item_formset'] = ActionItemFormSet(
                instance=self.incident,
                prefix='actionitems',
                form_kwargs={'incident': self.incident}
            )
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        action_item_formset = context['action_item_formset']

        if not action_item_formset.is_valid():
            messages.error(self.request, "There was an error with the action items. Please check the details.")
            return self.form_invalid(form)

        investigation = form.save(commit=False)
        investigation.incident = self.incident
        investigation.investigator = self.request.user
        investigation.completed_by = self.request.user
        investigation.personal_factors = json.loads(self.request.POST.get('personal_factors_json', '[]'))
        investigation.job_factors = json.loads(self.request.POST.get('job_factors_json', '[]'))
        investigation.save()

        action_items = action_item_formset.save(commit=False)
        has_forwarded_items = False

        for item in action_items:
            item.incident = self.incident
            item.created_by = self.request.user
            
            if item.assignment_type == 'SELF':
                item.status = 'COMPLETED'
                
                if not item.completion_date:
                    item.completion_date = timezone.now().date()
            else:
                item.status = 'PENDING'
                has_forwarded_items = True
            
            item.save()

        action_item_formset.save_m2m()

        for item in action_items:
            if item.assignment_type == 'SELF':
                item.responsible_person.add(self.request.user)
        
        for form_to_delete in action_item_formset.deleted_objects:
            form_to_delete.delete()
            
        try:
            from apps.notifications.services import NotificationService

            all_action_items = self.incident.action_items.all()  # ✅ FIXED

            for item in all_action_items:
                if item.assignment_type != 'SELF' and item.status == 'PENDING':
                    print("DEBUG: Sending notification for:", item.id)

                    NotificationService.notify(
                        content_object=item,
                        notification_type='INCIDENT_ACTION_ASSIGNED',
                        module='INCIDENT_ACTION'
                    )

        except Exception as e:
            print(f"Action item notification error: {e}")

        if has_forwarded_items:
            self.incident.status = 'ACTION_PLAN_PENDING'
        else:
            self.incident.status = 'PENDING_APPROVAL'
            
        self.incident.approval_status = 'PENDING'
        self.incident.investigation_completed_date = investigation.completed_date
        self.incident.save()

        # 4. Notifications
        try:
            from apps.notifications.services import NotificationService
            NotificationService.notify(
                content_object=investigation,
                notification_type='INCIDENT_INVESTIGATION_COMPLETED',
                module='INCIDENT_INVESTIGATION_REPORTED'
            )
            print("✅ Investigation completion notifications sent")
        except Exception as e:
            print(f"❌ Notification error: {e}")
            
            
            
        

        messages.success(self.request, "Investigation report submitted successfully.")
        return redirect('accidents:incident_detail', pk=self.incident.pk)
    
    def form_invalid(self, form):
        messages.error(self.request, "Please correct the errors in the form below.")
        context = self.get_context_data()
        action_item_formset = context['action_item_formset']
        if not action_item_formset.is_valid():
            # Yeh error reporting ko aur behtar banata hai
            for formset_form in action_item_formset:
                 for field, errors in formset_form.errors.items():
                      messages.error(self.request, f"Action Item Error in '{field}': {', '.join(errors)}")
        return super().form_invalid(form)
        
class ActionItemCreateView(LoginRequiredMixin, CreateView):
    """Create action item"""
    model = IncidentActionItem
    form_class = IncidentActionItemForm
    template_name = 'accidents/action_item_create.html'
    
    def dispatch(self, request, *args, **kwargs):
        self.incident = get_object_or_404(Incident, pk=self.kwargs['incident_pk'])
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['incident'] = self.incident
        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['incident'] = self.incident
        return kwargs

    def form_valid(self, form):
        # Link the action item to the incident
        form.instance.incident = self.incident
        response = super().form_valid(form)
        form.save_m2m()  # Save ManyToMany fields like responsible_person

        action_item = self.object

        # ===== ADD NOTIFICATION: Action Item Assigned =====
        try:
            from apps.notifications.services import NotificationService

            # Notify responsible persons + assigned_to
            extra_recipients = []
            if hasattr(action_item, 'assigned_to') and action_item.assigned_to:
                extra_recipients.append(action_item.assigned_to)

            NotificationService.notify(
                content_object=action_item,
                notification_type='INCIDENT_ACTION_ASSIGNED',
                module='INCIDENT_ACTION',
                extra_recipients=extra_recipients if extra_recipients else None
            )
            
            print("✅ Action assignment notifications sent")
        except Exception as e:
            print(f"❌ Error sending action assignment notifications: {e}")
        
        messages.success(self.request, 'Action item created successfully!')
        return response

    def get_success_url(self):
        return reverse_lazy(
            'accidents:incident_detail',
            kwargs={'pk': self.incident.pk}
        )
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['incident'] = self.incident
        return kwargs

 

class IncidentApprovalView(LoginRequiredMixin, DetailView):
    """
    Displays incident and investigation summaries for approval.
    Handles approving (which now closes the incident) or rejecting.
    """
    model = Incident
    template_name = 'accidents/incident_approval.html'
    context_object_name = 'incident'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['investigation_report'] = self.object.investigation_report
        return context

    def post(self, request, *args, **kwargs):
        """
        Handles 'approve' and 'reject' submissions according to the new status flow.
        """
        incident = self.get_object()
        
        if 'approve_action' in request.POST:
            # --- APPROVAL LOGIC ---
            # Approve hone par status 'Pending Close' hoga
            incident.status = 'PENDING_CLOSE'
            incident.approval_status = 'APPROVED'
            incident.approved_by = request.user
            incident.approved_date = timezone.now()
            incident.rejection_remarks = "" # Purane remarks saaf kar dein
            incident.save()
            messages.success(request, f"Incident {incident.report_number} has been approved and is now Pending Closure.")
            
        elif 'reject_action' in request.POST:
            # --- REJECTION LOGIC ---
            rejection_remarks = request.POST.get('rejection_remarks', '').strip()
            if not rejection_remarks:
                messages.error(request, "Rejection remarks are required to reject the report.")
                return redirect('accidents:incident_approve', pk=incident.pk)
            
            # Status ko 'REJECTED' set karein
            incident.status = 'REJECTED'
            incident.approval_status = 'REJECTED'
            incident.rejection_remarks = rejection_remarks
            
            # Approval data ko clear karein
            incident.approved_by = None
            incident.approved_date = None
            
            incident.save()
            
            messages.warning(request, f"Incident {incident.report_number} investigation has been REJECTED.")
            
        return redirect('accidents:incident_detail', pk=incident.pk)
        
# AJAX Views
class GetZonesForPlantAjaxView(LoginRequiredMixin, TemplateView):
    """AJAX view to get zones for selected plant"""
    
    def get(self, request, *args, **kwargs):
        plant_id = request.GET.get('plant_id')
        zones = Zone.objects.filter(plant_id=plant_id, is_active=True).values('id', 'name', 'code')
        return JsonResponse(list(zones), safe=False)


class GetLocationsForZoneAjaxView(LoginRequiredMixin, TemplateView):
    """AJAX view to get locations for selected zone"""
    
    def get(self, request, *args, **kwargs):
        zone_id = request.GET.get('zone_id')
        locations = Location.objects.filter(zone_id=zone_id, is_active=True).values('id', 'name', 'code')
        return JsonResponse(list(locations), safe=False)
    
class GetSublocationsForLocationAjaxView(LoginRequiredMixin, TemplateView):
    """AJAX view to get sublocations for selected location"""
    
    def get(self, request, *args, **kwargs):
        location_id = request.GET.get('location_id')
        from apps.organizations.models import SubLocation
        sublocations = SubLocation.objects.filter(
            location_id=location_id, 
            is_active=True
        ).values('id', 'name', 'code')
        return JsonResponse(list(sublocations), safe=False)

    
from django.views import View

class IncidentPDFDownloadView(LoginRequiredMixin, View):
    """Generate PDF report for incident"""
    
    def get(self, request, pk):
        incident = get_object_or_404(Incident, pk=pk)
        
        # Check permissions
        if not (request.user.is_superuser or 
                request.user == incident.reported_by or
                request.user.role.name in ['ADMIN', 'SAFETY MANAGER', 'PLANT HEAD']):
            messages.error(request, "You don't have permission to view this report")
            return redirect('accidents:incident_list')
        
        return generate_incident_pdf(incident)  
    




class IncidentAccidentDashboardView(LoginRequiredMixin, TemplateView):
    """Incident Management Dashboard with Analytics and Filters"""
    template_name = 'accidents/accidents_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        today = timezone.now().date() # Use timezone aware date
        user = self.request.user

        # ==================================================
        # GET FILTER PARAMETERS FROM REQUEST
        # ==================================================
        selected_plant = self.request.GET.get('plant', '')
        selected_zone = self.request.GET.get('zone', '')
        selected_location = self.request.GET.get('location', '')
        selected_sublocation = self.request.GET.get('sublocation', '')
        selected_month = self.request.GET.get('month', '')

        # ==================================================
        # BASE QUERYSET - Filter by user role
        # ==================================================
        # --- FIX IS HERE ---
        # The role check was `getattr(user, 'role', None) == 'ADMIN'`, which is incorrect.
        # It should check the `name` attribute of the role object, like this:
        if user.is_superuser or (hasattr(user, 'role') and user.role and user.role.name == 'ADMIN'):
            incidents = Incident.objects.all()
        elif getattr(user, 'plant', None):
            incidents = Incident.objects.filter(plant=user.plant)
        else:
            incidents = Incident.objects.filter(reported_by=user)

        # ==================================================
        # APPLY FILTERS TO QUERYSET
        # ==================================================
        if selected_plant:
            incidents = incidents.filter(plant_id=selected_plant)
        if selected_zone:
            incidents = incidents.filter(zone_id=selected_zone)
        if selected_location:
            incidents = incidents.filter(location_id=selected_location)
        if selected_sublocation:
            incidents = incidents.filter(sublocation_id=selected_sublocation)
        if selected_month:
            try:
                year, month = map(int, selected_month.split('-'))
                incidents = incidents.filter(
                    incident_date__year=year,
                    incident_date__month=month
                )
            except ValueError:
                pass

        # ==================================================
        # POPULATE FILTER DROPDOWNS
        # ==================================================
        # --- AND THE SAME FIX IS APPLIED HERE ---
        if user.is_superuser or (hasattr(user, 'role') and user.role and user.role.name == 'ADMIN'):
            all_plants = Plant.objects.filter(is_active=True).order_by('name')
        elif getattr(user, 'plant', None):
            all_plants = Plant.objects.filter(id=user.plant.id, is_active=True)
        else:
            all_plants = Plant.objects.none()
        
        context['plants'] = all_plants

        # This logic populates the dependent dropdowns and is now correct
        # because `all_plants` is populated correctly.
        if selected_plant:
            context['zones'] = Zone.objects.filter(plant_id=selected_plant, is_active=True).order_by('name')
            if selected_zone:
                context['locations'] = Location.objects.filter(zone_id=selected_zone, is_active=True).order_by('name')
                if selected_location:
                    context['sublocations'] = SubLocation.objects.filter(location_id=selected_location, is_active=True).order_by('name')
                else:
                    context['sublocations'] = SubLocation.objects.filter(location__zone_id=selected_zone, is_active=True).order_by('name')
            else:
                context['locations'] = Location.objects.filter(zone__plant_id=selected_plant, is_active=True).order_by('name')
                context['sublocations'] = SubLocation.objects.filter(location__zone__plant_id=selected_plant, is_active=True).order_by('name')
        else:
            # When no plant is selected, show zones/locations for all accessible plants
            context['zones'] = Zone.objects.filter(plant__in=all_plants, is_active=True).order_by('name')
            context['locations'] = Location.objects.filter(zone__plant__in=all_plants, is_active=True).order_by('name')
            context['sublocations'] = SubLocation.objects.filter(location__zone__plant__in=all_plants, is_active=True).order_by('name')

        # The rest of the method remains the same...
        # (month_options, selected filter values, stats, chart data, etc.)
        
        month_options = []
        for i in range(12):
            current_month = today.month - i
            current_year = today.year
            if current_month <= 0:
                current_month += 12
                current_year -= 1
            date = datetime.date(current_year, current_month, 1)
            month_options.append({
                'value': date.strftime('%Y-%m'),
                'label': date.strftime('%B %Y')
            })
        context['month_options'] = month_options

        context['selected_plant'] = selected_plant
        context['selected_zone'] = selected_zone
        context['selected_location'] = selected_location
        context['selected_sublocation'] = selected_sublocation
        context['selected_month'] = selected_month
        
        # Get names for active filter display
        context['selected_plant_name'] = Plant.objects.get(id=selected_plant).name if selected_plant else ''
        context['selected_zone_name'] = Zone.objects.get(id=selected_zone).name if selected_zone else ''
        context['selected_location_name'] = Location.objects.get(id=selected_location).name if selected_location else ''
        context['selected_sublocation_name'] = SubLocation.objects.get(id=selected_sublocation).name if selected_sublocation else ''
        if selected_month:
            try:
                year, month = map(int, selected_month.split('-'))
                context['selected_month_label'] = datetime.date(year, month, 1).strftime('%B %Y')
            except:
                context['selected_month_label'] = ''

        context['has_active_filters'] = bool(
            selected_plant or selected_zone or selected_location or 
            selected_sublocation or selected_month
        )

        context['total_incidents'] = incidents.count()
        context['open_incidents'] = incidents.exclude(status='CLOSED').count()

        if selected_month:
            try:
                year, month = map(int, selected_month.split('-'))
                context['this_month_incidents'] = incidents.filter(
                    incident_date__year=year, incident_date__month=month
                ).count()
                context['current_month_name'] = datetime.date(year, month, 1).strftime('%B')
                context['current_year'] = year
            except:
                pass
        else:
            context['this_month_incidents'] = incidents.filter(
                incident_date__month=today.month, incident_date__year=today.year
            ).count()
            context['current_month_name'] = today.strftime('%B')
            context['current_year'] = today.year
        
        context['investigation_pending'] = incidents.filter(
            investigation_required=True,
            investigation_completed_date__isnull=True
        ).count()

        # =================================================================
        # START: MODIFIED SECTION
        # =================================================================

        # REMOVE these hardcoded lines:
        # context['lti_count'] = incidents.filter(incident_type__code='LTI').count()
        # context['mtc_count'] = incidents.filter(incident_type__code='MTC').count()
        # context['fa_count']  = incidents.filter(incident_type__code='FA').count()
        # context['hlfi_count'] = incidents.filter(incident_type__code='HLFI').count()

        # ADD this dynamic query for the doughnut chart:
        type_distribution = incidents.values(
            'incident_type__name', 'incident_type__code'
        ).annotate(
            count=Count('id')
        ).order_by('-count')

        # Prepare data for Chart.js
        type_chart_labels = [item['incident_type__name'] for item in type_distribution if item['incident_type__name']]
        type_chart_data = [item['count'] for item in type_distribution if item['incident_type__name']]

        # Pass the dynamic data to the template context
        context['type_chart_labels'] = json.dumps(type_chart_labels)
        context['type_chart_data'] = json.dumps(type_chart_data)

        # =================================================================
        # END: MODIFIED SECTION
        # =================================================================

        context['recent_incidents'] = incidents.select_related(
            'plant', 'location', 'reported_by'
        ).order_by('-reported_date')[:10]
        context['overdue_investigations'] = incidents.filter(
            investigation_required=True,
            investigation_completed_date__isnull=True,
            investigation_deadline__lt=today
        )

        six_months_ago = today - datetime.timedelta(days=180)
        monthly_incidents = incidents.filter(
            incident_date__gte=six_months_ago
        ).annotate(
            month=TruncMonth('incident_date')
        ).values('month').annotate(
            count=Count('id')
        ).order_by('month')

        monthly_labels = [item['month'].strftime('%b %Y') for item in monthly_incidents]
        monthly_data = [item['count'] for item in monthly_incidents]
        
        context['monthly_labels'] = json.dumps(monthly_labels)
        context['monthly_data'] = json.dumps(monthly_data)

        incident_types = IncidentType.objects.order_by('id')
        type_counts = incidents.values('incident_type__id').annotate(count=Count('id'))
        count_map = {item['incident_type__id']: item['count'] for item in type_counts}
        severity_labels = [itype.name for itype in incident_types]
        severity_data = [count_map.get(itype.id, 0) for itype in incident_types]

        context['severity_labels'] = json.dumps(severity_labels)
        context['severity_data'] = json.dumps(severity_data)

        status_distribution = incidents.values('status').annotate(count=Count('id')).order_by('-count')
        status_choices_dict = dict(Incident.STATUS_CHOICES)
        status_labels = [status_choices_dict.get(item['status'], item['status']) for item in status_distribution]
        status_data = []
        for item in status_distribution:
            filter_params = {'status': item['status']}
            list_url = reverse('accidents:incident_list') + '?' + urlencode(filter_params)
            status_data.append({'count': item['count'], 'url': list_url})

        context['status_labels'] = json.dumps(status_labels)
        context['status_data'] = json.dumps(status_data)

        last_month_start = (today.replace(day=1) - datetime.timedelta(days=1)).replace(day=1)
        last_month_count = incidents.filter(
            incident_date__year=last_month_start.year,
            incident_date__month=last_month_start.month
        ).count()

        current_month_count = context.get('this_month_incidents', 0)
        change = round(((current_month_count - last_month_count) / last_month_count) * 100, 1) if last_month_count > 0 else 0
        context['total_incidents_change'] = change
        context['total_incidents_change_abs'] = abs(change)

        return context

######################Closure 

class IncidentClosureCheckView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    Pre-closure verification page.
    Also handles the attachment upload directly on this page.
    """
    template_name = 'accidents/incident_closure_check.html'
    
    def test_func(self):
        """Check if user has permission to view this page."""
        return (
            self.request.user.is_superuser or
            # self.request.user.can_close_incidents or # Uncomment if you have this on your user model
            self.request.user.role.name in ['ADMIN', 'SAFETY MANAGER', 'PLANT HEAD']
        )
    
    def get_context_data(self, **kwargs):
        """Helper method to gather all context data."""
        context = {}
        incident = get_object_or_404(Incident, pk=self.kwargs['pk'])
        
        # --- THIS IS THE CORRECTED LINE ---
        # Removed the parentheses from incident.can_be_closed
        can_close, message = incident.can_be_closed
        # ------------------------------------
        
        try:
            investigation = incident.investigation_report
        except IncidentInvestigationReport.DoesNotExist:
            investigation = None
        
        action_items = incident.action_items.all()
        pending_actions = action_items.exclude(status='COMPLETED')
        completed_actions = action_items.filter(status='COMPLETED')
        
        context.update({
            'incident': incident,
            'can_close': can_close,
            'closure_message': message,
            'investigation': investigation,
            'action_items': action_items,
            'pending_actions': pending_actions,
            'completed_actions': completed_actions,
            'total_actions': action_items.count(),
            'completed_count': completed_actions.count(),
        })
        
        if 'attachment_form' not in kwargs:
             context['attachment_form'] = IncidentAttachmentForm(instance=incident)

        return context

    def get(self, request, *args, **kwargs):
        """Handles the display of the verification page."""
        context = self.get_context_data()
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        """Handles the file upload form submission."""
        incident = get_object_or_404(Incident, pk=self.kwargs['pk'])
        form = IncidentAttachmentForm(request.POST, request.FILES, instance=incident)

        if form.is_valid():
            form.save()
            messages.success(request, 'Closure attachment has been successfully uploaded.')
            return redirect('accidents:incident_closure_check', pk=incident.pk)
        else:
            messages.error(request, 'There was an error uploading the file. Please try again.')
            context = self.get_context_data(attachment_form=form)
            return render(request, self.template_name, context)
        

class IncidentClosureView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """Close an incident"""
    model = Incident
    form_class = IncidentClosureForm
    template_name = 'accidents/incident_closure.html'
    
    def test_func(self):
        """Check if user has permission to close incidents"""
        return (
            self.request.user.is_superuser or
            getattr(self.request.user, 'can_close_incidents', False) or
            self.request.user.role.name in ['ADMIN', 'SAFETY MANAGER', 'PLANT HEAD']
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        incident = self.object
        
        # Verify closure eligibility
        can_close, message = incident.can_be_closed

        context.update({
            'incident': incident,
            'can_close': can_close,
            'closure_message': message,
        })
        
        return context
    
    def form_valid(self, form):

        # Verify incident can be closed
        incident = form.save(commit=False)

        can_close, message = incident.can_be_closed
        if not can_close:
            messages.error(self.request, f"Cannot close incident: {message}")
            return redirect('accidents:incident_detail', pk=incident.pk)

        # ✅ Set closure fields
        incident.status = 'CLOSED'
        incident.closure_date = timezone.now()
        incident.closed_by = self.request.user
        incident.save()

        # print(f"Incident {incident.report_number} closed. Sending notification...")

        # For sending notifications
        try:
            from apps.notifications.services import NotificationService
            NotificationService.notify(
                content_object=incident,
                notification_type='INCIDENT_CLOSED',
                module='INCIDENT_CLOSED'
            )
            print("✅ Incident closure notifications sent")
        except Exception as e:
            print(f"❌ Error sending closure notifications: {e}")

        messages.success(
            self.request,
            f'Incident {incident.report_number} has been successfully closed.'
        )

        return redirect('accidents:incident_detail', pk=incident.pk)
    
    def get_success_url(self):
        return reverse_lazy('accidents:incident_detail', kwargs={'pk': self.object.pk})
 

class IncidentReopenView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Reopen a closed incident"""
    
    def test_func(self):
        return (
            self.request.user.is_superuser or
            self.request.user.role.name in ['ADMIN', 'SAFETY MANAGER']
        )
    
    def post(self, request, pk):
        incident = get_object_or_404(Incident, pk=pk)
        
        if incident.status != 'CLOSED':
            messages.error(request, "Only closed incidents can be reopened")
            return redirect('accidents:incident_detail', pk=pk)
        
        # Reopen incident
        incident.status = 'UNDER_INVESTIGATION'
        incident.closure_date = None
        incident.closed_by = None
        incident.save()
        
        messages.warning(
            request,
            f'Incident {incident.report_number} has been reopened for further investigation.'
        )
        
        return redirect('accidents:incident_detail', pk=pk)    
    
class InvestigationDetailView(LoginRequiredMixin, DetailView):
    """View investigation report details"""
    model = IncidentInvestigationReport
    template_name = 'accidents/investigation_detail.html'
    context_object_name = 'investigation'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['incident'] = self.object.incident
        context['cancel_url'] = (self.request.GET.get('next') or self.request.META.get('HTTP_REFERER') or '/')
        
        return context    
    


##################notification 
class NotificationListView(LoginRequiredMixin, ListView):
    """List user's notifications"""
    model = IncidentNotification
    template_name = 'accidents/notifications.html'
    context_object_name = 'notifications'
    paginate_by = 20
    
    def get_queryset(self):
        return IncidentNotification.objects.filter(
            recipient=self.request.user
        ).select_related('incident', 'incident__plant', 'incident__location')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['unread_count'] = self.get_queryset().filter(is_read=False).count()
        return context


class MarkNotificationReadView(LoginRequiredMixin, View):
    """Mark notification as read"""
    
    def post(self, request, pk):
        notification = get_object_or_404(
            IncidentNotification, 
            pk=pk, 
            recipient=request.user
        )
        notification.mark_as_read()
        return JsonResponse({'status': 'success'})
#########################################33

class MarkAllNotificationsReadView(LoginRequiredMixin, View):
    """Mark all notifications as read"""
    
    def post(self, request):
        IncidentNotification.objects.filter(
            recipient=request.user,
            is_read=False
        ).update(is_read=True, read_at=timezone.now())
        return JsonResponse({'status': 'success'})   




class IncidentFilterMixin:
    """
    A mixin to provide a filtered queryset of incidents based on URL parameters.
    This can be reused by the Dashboard, Export views, etc.
    """
    def get_filtered_queryset(self):
        user = self.request.user
        
        # Get filter parameters from the URL
        selected_plant = self.request.GET.get('plant', '')
        selected_zone = self.request.GET.get('zone', '')
        selected_location = self.request.GET.get('location', '')
        selected_sublocation = self.request.GET.get('sublocation', '')
        selected_month = self.request.GET.get('month', '')
        selected_type = self.request.GET.get('type', '')
        selected_status = self.request.GET.get('status', '')

        # Base queryset based on user's role
        if user.is_superuser or getattr(user, 'role', None) == 'ADMIN':
            base_incidents = Incident.objects.select_related(
                'plant', 'zone', 'location', 'sublocation', 'reported_by', 'closed_by'
            ).all()
        elif getattr(user, 'plant', None):
            base_incidents = Incident.objects.filter(plant=user.plant).select_related(
                'plant', 'zone', 'location', 'sublocation', 'reported_by', 'closed_by'
            )
        else:
            base_incidents = Incident.objects.filter(reported_by=user).select_related(
                'plant', 'zone', 'location', 'sublocation', 'reported_by', 'closed_by'
            )

        # Apply filters
        incidents = base_incidents
        if selected_plant:
            incidents = incidents.filter(plant_id=selected_plant)
        if selected_zone:
            incidents = incidents.filter(zone_id=selected_zone)
        if selected_location:
            incidents = incidents.filter(location_id=selected_location)
        if selected_sublocation:
            incidents = incidents.filter(sublocation_id=selected_sublocation)
        if selected_type:
            incidents = incidents.filter(incident_type=selected_type)
        if selected_status == 'open':
            incidents = incidents.exclude(status='CLOSED')
        elif selected_status:
            incidents = incidents.filter(status=selected_status)

        if selected_month:
            try:
                year, month = map(int, selected_month.split('-'))
                incidents = incidents.filter(incident_date__year=year, incident_date__month=month)
            except (ValueError, TypeError):
                pass
        
        return incidents.order_by('-incident_date', '-incident_time')
    

class ExportIncidentsExcelView(LoginRequiredMixin, IncidentFilterMixin, View):
    def get(self, request, *args, **kwargs):
        user = self.request.user
        
        # 1. Sabse pehle Base Queryset set karein (Permission ke hisaab se)
        if user.is_superuser or (hasattr(user, 'role') and user.role and user.role.name == 'ADMIN'):
            # Admin sabka data dekh sakta hai
            queryset = Incident.objects.all()
        elif getattr(user, 'plant', None):
            # Plant user sirf apne assigned plant ka data dekh sakta hai
            queryset = Incident.objects.filter(plant=user.plant)
        else:
            # Baki log sirf apna reported data dekh sakte hain
            queryset = Incident.objects.filter(reported_by=user)

        # 2. Ab URL filters apply karein (Plant, Zone, Month etc.)
        # Hum IncidentFilterMixin ka direct use nahi karenge balki manually filter karenge 
        # taaki koi user URL modify karke dusre plant ka data na le sake
        
        selected_plant = request.GET.get('plant')
        selected_zone = request.GET.get('zone')
        selected_location = request.GET.get('location')
        selected_sublocation = request.GET.get('sublocation')
        selected_month = request.GET.get('month')

        # Admin agar kisi specific plant ko filter kare
        if selected_plant and (user.is_superuser or (hasattr(user, 'role') and user.role.name == 'ADMIN')):
            queryset = queryset.filter(plant_id=selected_plant)
        
        if selected_zone:
            queryset = queryset.filter(zone_id=selected_zone)
        if selected_location:
            queryset = queryset.filter(location_id=selected_location)
        if selected_sublocation:
            queryset = queryset.filter(sublocation_id=selected_sublocation)
        
        if selected_month:
            try:
                year, month = map(int, selected_month.split('-'))
                queryset = queryset.filter(incident_date__year=year, incident_date__month=month)
            except ValueError:
                pass

        # Optimize for Excel export
        queryset = queryset.select_related('incident_type', 'plant', 'zone', 'location', 'sublocation', 'reported_by', 'closed_by')

        # --- Baaki ka Excel generation code yahan se start hoga ---
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Incident Report'

        # --- Define Styles ---
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        row_fills = [
            PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"),
            PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        ]

        status_fills = {
            'Open': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            'In Progress': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            'Closed': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        }

        # New style for wrapping text in specific columns
        wrap_alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

        # --- Headers ---
        headers = [
            'Report Number', 'Incident Type', 'Status', 'Incident Date', 'Incident Time',
            'Plant', 'Zone', 'Location', 'Sub-Location', 'Description', 'Affected Person',
            'Nature of Injury', 'Reported By', 'Reported Date', 'Investigation Deadline',
            'Closure Date', 'Closed By'
        ]
        sheet.append(headers)

        # Style the header row
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # --- Data Population and Styling ---
        # Get column index for text wrapping before the loop
        desc_col_idx = headers.index('Description') + 1
        injury_col_idx = headers.index('Nature of Injury') + 1
        
        for row_index, incident in enumerate(queryset, start=2):
            row_data = [
                incident.report_number,
                incident.incident_type.name if incident.incident_type else 'N/A',
                incident.get_status_display(),
                incident.incident_date,
                incident.incident_time,
                incident.plant.name if incident.plant else 'N/A',
                incident.zone.name if incident.zone else 'N/A',
                incident.location.name if incident.location else 'N/A',
                incident.sublocation.name if incident.sublocation else 'N/A',
                incident.description,
                incident.affected_person_name,
                incident.nature_of_injury,
                incident.reported_by.get_full_name() if incident.reported_by else 'N/A',
                incident.reported_date.strftime("%Y-%m-%d %H:%M") if incident.reported_date else None,
                incident.investigation_deadline,
                incident.closure_date.strftime("%Y-%m-%d %H:%M") if incident.closure_date else None,
                incident.closed_by.get_full_name() if incident.closed_by else 'N/A'
            ]
            sheet.append(row_data)

            # Apply alternating row color (zebra striping)
            current_fill = row_fills[(row_index - 2) % 2]
            for cell in sheet[row_index]:
                cell.fill = current_fill
            
            # Apply wrap text alignment to specific cells
            sheet.cell(row=row_index, column=desc_col_idx).alignment = wrap_alignment
            sheet.cell(row=row_index, column=injury_col_idx).alignment = wrap_alignment

        # --- Conditional Formatting for Status ---
        if sheet.max_row >= 2:
            status_column_letter = get_column_letter(headers.index('Status') + 1)
            for status, fill in status_fills.items():
                rule = CellIsRule(operator='equal', formula=[f'"{status}"'], fill=fill)
                sheet.conditional_formatting.add(f'{status_column_letter}2:{status_column_letter}{sheet.max_row}', rule)

        # --- Adjust Column Widths ---
        column_widths = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    column_widths[cell.column_letter] = max(
                        (column_widths.get(cell.column_letter, 0), len(str(cell.value)))
                    )

        for col_letter, width in column_widths.items():
            header_name = sheet[f'{col_letter}1'].value
            # Set a fixed width for columns that need text wrapping
            if header_name in ['Description', 'Nature of Injury']:
                sheet.column_dimensions[col_letter].width = 50
            else:
                # Auto-size other columns with a little padding
                sheet.column_dimensions[col_letter].width = width + 2

        # --- HTTP Response ---
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f"Incident_Report_{timezone.now().strftime('%Y-%m-%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        workbook.save(response)

        return response
    

# class IncidentCloseView(LoginRequiredMixin, UpdateView):
#     """Close an incident"""
#     model = Incident
#     template_name = 'accidents/incident_close.html'
#     fields = ['closure_remarks', 'lessons_learned', 'preventive_measures']
#     def form_valid(self, form):
#         print(">>>> form_valid called")
#         incident = form.save(commit=False)

#         # Set closure fields
#         incident.status = 'CLOSED'
#         incident.closure_date = timezone.now()
#         incident.closed_by = self.request.user
#         incident.save()
#         print(f">>>> Incident {incident.report_number} closed, calling NotificationService")
#         # ===== ADD NOTIFICATION: Incident Closed =====
#         try:
#             from apps.notifications.services import NotificationService
#             NotificationService.notify(
#                 content_object=incident,
#                 notification_type='INCIDENT_CLOSED',
#                 module='INCIDENT_CLOSED'
#             )
#             print("✅ Incident closure notifications sent")

#         except Exception as e:
#             print(f"❌ Error sending closure notifications: {e}")

#         messages.success(self.request,f'Incident {incident.report_number} has been closed successfully.')
#         return redirect('accidents:incident_detail', pk=incident.pk)


class IncidentActionItemCompleteView(LoginRequiredMixin, FormView): # Cambiado de UpdateView a FormView
    """
    Permite a un usuario asignado marcar un elemento de acción de incidente como completo
    creando un registro de finalización.
    """
    form_class = IncidentActionItemCompleteForm
    template_name = 'accidents/action_item_complete.html'
    
    def setup(self, request, *args, **kwargs):
        """Obtiene el objeto de elemento de acción antes."""
        super().setup(request, *args, **kwargs)
        self.action_item = get_object_or_404(IncidentActionItem, pk=self.kwargs['pk'])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['action_item'] = self.action_item
        context['incident'] = self.action_item.incident
        if hasattr(self.action_item.incident, 'investigation_report'):
            context['investigation'] = self.action_item.incident.investigation_report
        else:
            context['investigation'] = None
        return context

    def form_valid(self, form):
        """
        - Crea un registro de ActionItemCompletion.
        - Verifica si TODOS los usuarios responsables han completado.
        - Si es así, actualiza el estado del elemento de acción principal a 'COMPLETED'.
        """
        # Crear la nueva entrada de finalización
        completion = form.save(commit=False)
        completion.action_item = self.action_item
        completion.completed_by = self.request.user
        completion.save()

        # Comprobar si todas las personas han completado
        assigned_users_count = self.action_item.responsible_person.count()
        completing_users_count = self.action_item.completions.count()

        if assigned_users_count > 0 and assigned_users_count == completing_users_count:
            # Todos han completado, así que actualiza el elemento de acción principal
            self.action_item.status = 'COMPLETED'
            self.action_item.save()
            messages.success(self.request, f"Action item for incident '{self.action_item.incident.report_number}' is now fully completed by all responsible persons.")
            
            # (Lógica existente para verificar si el incidente se puede mover a PENDING_APPROVAL)
            incident = self.action_item.incident
            if not incident.action_items.exclude(status='COMPLETED').exists():
                incident.status = 'PENDING_APPROVAL'
                incident.save()
                messages.info(self.request, f"All action items for incident {incident.report_number} are complete. The incident is now pending final approval.")
        else:
            remaining_count = assigned_users_count - completing_users_count
            messages.info(self.request, f"Your completion for the action item has been recorded. Still waiting for {remaining_count} other person(s) to complete.")

        return redirect(self.get_success_url())

    def get_success_url(self):
        return reverse_lazy('accidents:my_action_items')



class MyActionItemsView(LoginRequiredMixin, ListView):
    model = IncidentActionItem
    template_name = 'accidents/my_action_items.html'
    context_object_name = 'action_items'
    paginate_by = 15

    def get_queryset(self):
        user = self.request.user
        
        # Subconsulta: Verificar si existe un registro de finalización para el usuario actual
        user_completed_subquery = ActionItemCompletion.objects.filter(
            action_item=OuterRef('pk'),
            completed_by=user
        )

        queryset = IncidentActionItem.objects.filter(
            responsible_person=user
        ).annotate(
            is_done_by_me=Exists(user_completed_subquery) # Lógica de anotación actualizada
        ).select_related( 
            'incident', 
            'incident__plant',
            'created_by'
        ).order_by('target_date')

        # ... el resto de su lógica de filtro de estado ...
        status_filter = self.request.GET.get('status', '')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
            
        return queryset

    # ... get_context_data permanece igual ...
    def get_context_data(self, **kwargs):
        """
        Añade estadísticas y opciones de filtro al contexto de la plantilla.
        """
        context = super().get_context_data(**kwargs)
        user = self.request.user

        all_my_items = IncidentActionItem.objects.filter(responsible_person=user)

        context['total_assigned'] = all_my_items.count()
        context['pending_count'] = all_my_items.filter(status__in=['PENDING', 'IN_PROGRESS']).count()
        
        overdue_items = [item for item in all_my_items if item.is_overdue]
        context['overdue_count'] = len(overdue_items)

        context['status_choices'] = IncidentActionItem.STATUS_CHOICES
        context['selected_status'] = self.request.GET.get('status', '')

        return context